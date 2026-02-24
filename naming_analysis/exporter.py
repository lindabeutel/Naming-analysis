"""
exporter.py

Excel export module for the naming-analysis pipeline.

This module contains routines to produce a consolidated export workbook from the
project’s JSON/Excel data sources. It supports:

- Updating existing worksheets with confirmed naming variants.
- Updating collocation fields in existing worksheets.
- Creating additional export worksheets (e.g., categorized lemmas).
- Writing a final export Excel file for downstream analysis and archiving.

Formatting:
The export process aims to preserve or replicate workbook formatting using
openpyxl utilities where applicable.

Scope:
This module performs filesystem and workbook side effects (reading/writing Excel
files and modifying worksheets). It does not perform analytical computations
beyond the transformations required for export.
"""
# Standard library
import os
import shutil
from copy import copy

# Third-party libraries
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border
from openpyxl.utils import get_column_letter

# Internal project imports
from naming_analysis.shared import ask_user_choice, parse_verse_number
from naming_analysis.io_utils import safe_read_json

# ============================================================================
# Top-level export orchestration
# ============================================================================

def export_all_data_to_new_excel(book_name, paths, options):
    """
    Create a consolidated export workbook for a given corpus.

    This function copies the source Excel workbook and applies export updates
    based on the current session state:

    - Insert confirmed naming variants into the main worksheet ("Gesamt").
    - Update collocation fields in the main worksheet ("Gesamt").
    - Create an additional worksheet containing categorized lemmata.

    Side effects:
        - Creates/overwrites an export file at data/<book_name>/<book_name>_final.xlsx.
        - Reads the source Excel workbook and writes the exported workbook.
        - Prompts the user on PermissionError (locked file) and may return early.
        - Optionally opens the exported workbook via os.startfile (Windows only).

    Parameters:
        book_name (str):
            Corpus identifier used for output folder and filename.

        paths (dict):
            Session path dictionary. Expects at least:
                - "original_excel": source workbook path to copy from
                - "missing_naming_variants_json": naming variants container (JSON)
                - "collocations_json": collocation data (JSON)
                - "categorization_json": categorization data (JSON)

        options (dict):
            Boolean feature flags controlling which export steps are executed.
            Expected keys (German, aligned with Excel/export layer):
                - "benennungen"
                - "kollokationen"
                - "kategorisierung"

    Notes:
        The function also supports legacy/alternate keys:
        "json_benennungen", "json_kollokationen", "json_kategorisierung".
        If those keys are missing, it falls back to the corresponding
        canonical keys listed above.
    """
    print("Starting export of all naming variant data...")

    # Support alternate keys for JSON export paths (legacy naming)
    paths = {
        **paths,
        "json_benennungen": paths.get("json_benennungen") or paths.get("missing_naming_variants_json"),
        "json_kollokationen": paths.get("json_kollokationen") or paths.get("collocations_json"),
        "json_kategorisierung": paths.get("json_kategorisierung") or paths.get("categorization_json"),
    }

    # Ensure corpus-specific output directory exists
    project_dir = os.path.join("data", book_name)
    os.makedirs(project_dir, exist_ok=True)
    target_path = os.path.join(project_dir, f"{book_name}_final.xlsx")

    # Copy the source Excel workbook (retry on PermissionError / locked file)
    while True:
        try:
            shutil.copy(paths["original_excel"], target_path)
            break
        except PermissionError:
            print("The Excel file is currently open or locked.")
            print("Please close the file and try again.")
            retry = ask_user_choice("Retry export? (y/n): ", ["y", "n"])
            if retry != "y":
                return

    wb = openpyxl.load_workbook(target_path)
    sheet = wb["Gesamt"]

    # Update existing "Gesamt" worksheet with selected export steps
    if options.get("benennungen", False):
        print("Exporting confirmed naming variants...")
        insert_naming_variants(sheet, paths["json_benennungen"])

    if options.get("kollokationen", False):
        print("Exporting collocations...")
        update_collocations(sheet, paths["json_kollokationen"])

    # Create additional worksheet for categorized lemma (if selected)
    if options.get("kategorisierung", False):
        print("Exporting categorized lemma (this may take a second)...")
        create_categorized_lemmas_sheet(wb, sheet, paths["json_kategorisierung"])

    wb.save(target_path)
    print(f"Export completed: {target_path}")

    # Optional open (Windows-only)
    answer = ask_user_choice(
        f"Do you want to open the Excel file '{os.path.basename(target_path)}' now? (y/n):", ["y", "n"]).strip().lower()
    if answer == "y":
        try:
            os.startfile(os.path.abspath(target_path))  # Windows only
        except Exception as e:
            print(f"Could not open file: {e}")

# ============================================================================
# Formatting utilities (openpyxl-related)
# ============================================================================

def get_format_template(sheet, column_index):
    """
    Extract the formatting style from the first non-empty, styled cell
    in the specified column (starting from row 2).

    The function scans the column top-down and returns copies of the
    cell's formatting attributes to allow safe reuse without mutating
    the original workbook styles.

    Parameters:
        sheet (Worksheet):
            Target worksheet (openpyxl).
        column_index (int):
            1-based column index to inspect.

    Returns:
        tuple:
            (font, alignment, border, number_format)

            Each element may be None if no styled cell was found.
            If no styled, non-empty cell exists in the column,
            (None, None, None, None) is returned.

    Notes:
        - Row 1 is skipped (assumed header row).
        - Only the first styled, non-empty cell is used as template.
        - Returned style objects are shallow copies.
    """
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=column_index)
        if cell.value:
            if cell.has_style:
                return copy(cell.font), copy(cell.alignment), copy(cell.border), cell.number_format
    return None, None, None, None

# ============================================================================
# Sheet update helpers (modify existing worksheets)
# ============================================================================

def insert_naming_variants(sheet, json_path: str) -> None:
    """
    Append confirmed naming variant entries from a JSON container into the Excel worksheet.

    For each JSON entry with Status == "confirmed", this function appends a new row
    to the target worksheet (typically "Gesamt"). Appended rows are highlighted and
    formatted to match the existing column style templates.

    Formatting behavior:
        - Column styles (font, alignment, border, number_format) are copied from the
          first styled, non-empty cell found in each column (via get_format_template()).
        - The appended row is visually highlighted via a solid fill color.
        - For the verse column ("Vers", column 2), numeric values are formatted as:
            - integers: "0"
            - decimals: "0.00"

    Parameters:
        sheet (Worksheet):
            Target worksheet (expected to be the main sheet, e.g., "Gesamt").
        json_path (str):
            Path to the naming variants JSON file.

    Notes (Beta state):
        - No schema validation of JSON content is performed beyond key lookups.
        - Column order is assumed to match the expected export layout.
        - Existing worksheet data is not deduplicated against inserted entries.
    """
    data = safe_read_json(json_path, default=[])

    confirmed_entries = [entry for entry in data if entry.get("Status") == "confirmed"]
    if not confirmed_entries:
        print("No confirmed naming variants to insert.")
        return

    last_line = sheet.max_row + 1
    fill_color = PatternFill(
        start_color="F9C691",
        end_color="F9C691",
        fill_type="solid"
    )

    for entry in confirmed_entries:
        new_line = [
            entry.get("Benannte Figur", ""),
            entry.get("Vers", ""),
            entry.get("Eigennennung", ""),
            entry.get("Nennende Figur", ""),
            entry.get("Bezeichnung", ""),
            entry.get("Erzähler", ""),
            entry.get("Kollokation", "")
        ]

        for col_num, value in enumerate(new_line, start=1):
            cell = sheet.cell(row=last_line, column=col_num, value=value)

            font_tpl, alignment_tpl, border_tpl, number_format_tpl = get_format_template(sheet, col_num)
            if font_tpl:
                cell.font = font_tpl
                cell.alignment = alignment_tpl
                cell.border = border_tpl
                # Format "Vers" column (column 2): integers as "0", decimals as "0.00"
                if col_num == 2 and isinstance(value, (int, float)):
                    if value % 1 == 0:
                        cell.number_format = "0"
                    else:
                        cell.number_format = "0.00"
                else:
                    cell.number_format = number_format_tpl

            cell.fill = fill_color

        last_line += 1

    print("Naming variants successfully added.")

def update_collocations(sheet, json_path: str) -> None:
    """
    Update the "Kollokationen" column in an Excel worksheet using collocation JSON data.

    This function maps worksheet rows by verse number and writes the collocation string
    to the corresponding "Kollokationen" cells. Column formatting is preserved by copying
    a template style from the first styled, non-empty cell in the collocation column.

    Matching behavior:
        - Worksheet rows are indexed by the value in the "Vers" column.
        - Verse values from both the worksheet and the JSON entries are normalized
          via parse_verse_number() for matching (supports decimals and Excel-style formats).
        - For each JSON entry, all worksheet rows with the same normalized verse number are updated.

    Parameters:
        sheet (Worksheet):
            Target worksheet to update (openpyxl).
        json_path (str):
            Path to the collocations JSON file.

    Notes (Beta state):
        - The function assumes the header row contains "Vers" and "Kollokationen".
        - No schema validation of JSON entries is performed (expects keys "Vers" and "Kollokationen").
        - Verse matching relies on project-standard parsing (parse_verse_number) rather than int(...).
    """
    data = safe_read_json(json_path, default=[])

    header = [cell.value for cell in sheet[1]]
    try:
        verse_col = header.index("Vers") + 1
        collocation_col = header.index("Kollokationen") + 1
    except ValueError:
        print("Columns 'Vers' or 'Kollokationen' not found!")
        return

    # Build an index: verse_number -> list of worksheet row indices
    verse_to_rows = {}
    for row in range(2, sheet.max_row + 1):
        verse_value = sheet.cell(row=row, column=verse_col).value
        verse_key = parse_verse_number(verse_value, fallback=-1)
        if verse_key != -1:
            verse_to_rows.setdefault(verse_key, []).append(row)

    font_tpl, alignment_tpl, border_tpl, number_format_tpl = get_format_template(sheet, collocation_col)

    updated_count = 0
    for entry in data:
        verse = parse_verse_number(entry["Vers"], fallback=-1)
        if verse == -1:
            continue
        new_value = entry["Kollokationen"]

        # Update all rows that match the verse number
        matching_rows = verse_to_rows.get(verse, [])
        for row in matching_rows:
            cell = sheet.cell(row=row, column=collocation_col, value=new_value)
            if font_tpl:
                cell.font = font_tpl
                cell.alignment = alignment_tpl
                cell.border = border_tpl
                cell.number_format = number_format_tpl
            updated_count += 1

    print(f"{updated_count} collocations successfully updated.")

# ============================================================================
# Sheet generation helper (create new worksheets)
# ============================================================================

def create_categorized_lemmas_sheet(wb, _, json_path: str) -> None:
    """
    Create or replace the worksheet "lemmatisiert" containing categorized entries.

    The function deletes an existing sheet with the same name (if present),
    creates a new worksheet, and writes structured annotation data from JSON
    into predefined columns.

    Formatting:
        - Header row is written in bold.
        - Data rows use regular formatting.
        - Column widths are standardized.
        - The first row is frozen and auto-filter is enabled.
        - The "Vers" column is formatted as "0" or "0.00" for numeric values.

    Parameters:
        wb (Workbook):
            Target Excel workbook (openpyxl).
        _:
            Placeholder parameter for interface compatibility
            (original sheet not required here).
        json_path (str):
            Path to the categorized entries JSON file.

    Side effects:
        - May delete an existing worksheet named "lemmatisiert".
        - Modifies workbook structure and formatting.
        - Reorders workbook sheets.

    Notes (Beta state):
        - Sheet reordering uses the internal wb._sheets list because wb.move_sheet()
          was not sufficient/reliable for the desired placement in this project context.
    """
    # --- Load JSON annotation data ---
    # Expected: list of dicts with categorized naming data
    annotations = safe_read_json(json_path, default=[])

    # --- Replace existing worksheet if it already exists ---
    # Ensures deterministic regeneration of the sheet
    if "lemmatisiert" in wb.sheetnames:
        del wb["lemmatisiert"]

    # Create new worksheet at end (position adjusted later)
    ws_new = wb.create_sheet("lemmatisiert")

    # --- Define base cell styles ---
    # Font styles (header bold, data regular)
    regular_font = Font(name="Times New Roman", size=8, bold=False)
    bold_font = Font(name="Times New Roman", size=8, bold=True)

    # Left-aligned, bottom vertical alignment (project convention)
    default_alignment = Alignment(horizontal="left", vertical="bottom")

    # Empty/default border (no explicit border styling applied)
    default_border = Border()

    # --- Define export column structure ---
    # Explicit ordering ensures deterministic column layout
    headers = [
        "Benannte Figur", "Vers", "Eigennennung", "Nennende Figur", "Bezeichnung", "Erzähler",
        "Bezeichnung 1", "Bezeichnung 2", "Bezeichnung 3", "Bezeichnung 4",
        "Epitheta 1", "Epitheta 2", "Epitheta 3", "Epitheta 4", "Epitheta 5"
    ]

    # Convert JSON list into DataFrame for structured column handling
    df = pd.DataFrame(annotations)

    # Ensure all expected columns exist (missing columns filled with empty string)
    for col in headers:
        if col not in df.columns:
            df[col] = ""

    # Reorder DataFrame to match defined header order
    df = df[headers]

    # --- Write header row ---
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)

        # Write header label
        cell = ws_new.cell(row=1, column=col_idx, value=header)

        # Standardize column width for readability
        ws_new.column_dimensions[col_letter].width = 20

        # Apply header formatting
        cell.font = bold_font
        cell.alignment = default_alignment
        cell.border = default_border
        cell.number_format = "General"

    # --- Write data rows ---
    # Iterate over DataFrame rows and write structured annotation data
    # starting from Excel row 2 (row 1 contains headers)
    for row_idx, row in df.iterrows():
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_new.cell(row=row_idx + 2, column=col_idx, value=row[header])

            # Apply base formatting (consistent across all data cells)
            cell.font = regular_font
            cell.alignment = default_alignment
            cell.border = default_border

            # Special formatting rule for verse column:
            # integers → "0"
            # decimals → "0.00"
            if header == "Vers" and isinstance(row[header], (int, float)):
                if row[header] % 1 == 0:
                    cell.number_format = "0"
                else:
                    cell.number_format = "0.00"
            else:
                cell.number_format = "General"

    # --- Freeze header row ---
    # Keeps column labels visible during vertical scrolling
    ws_new.freeze_panes = "A2"

    # --- Enable auto-filter ---
    # Applies filter dropdowns to header row across all defined columns
    ws_new.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # --- Reorder worksheet ---
    # Move the new sheet to second position in workbook.
    # Uses internal wb._sheets because wb.move_sheet() was not sufficient/reliable here.
    wb._sheets.insert(1, wb._sheets.pop(wb._sheets.index(ws_new)))

    print("Worksheet 'lemmatisiert' successfully created.")