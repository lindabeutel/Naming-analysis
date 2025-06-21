import tkinter as tk
from tkinter import filedialog

import pandas as pd
import re
import json
import os
import shutil

import xml.etree.ElementTree as ET
from xml.etree.ElementTree import Element

from copy import copy, deepcopy

from typing import Union

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

DataType = dict[str, Union[pd.DataFrame, Element, str, None]]
tei_ns = {'tei': 'http://www.tei-c.org/ns/1.0'}

def initialize_project():
    book_name = input("Which book are we working on today? (e.g., Eneasroman): ").strip()
    book_name = book_name[0].upper() + book_name[1:]

    os.makedirs("data", exist_ok=True)

    config_path = os.path.join("data", f"config_{book_name}.json")
    progress_path = os.path.join("data", f"progress_{book_name}.json")

    paths = {
        "missing_namings_json": os.path.join("data", f"missing_namings_{book_name}.json"),
        "progress_json": progress_path,
        "collocations_json": os.path.join("data", f"collocations_{book_name}.json"),
        "categorization_json": os.path.join("data", f"categorization_{book_name}.json"),
        "lemma_normalization_json": os.path.join("data", "lemma_normalization.json"),
        "ignored_lemmas_json": os.path.join("data", "ignored_lemmas.json"),
        "lemma_categories_json": os.path.join("data", "lemma_categories.json"),
        "config_json": config_path
    }

    # Fortschritt laden oder initialisieren
    namings_last_verse = 0
    collocations_last_verse = 0
    categorization_last_verse = 0

    if os.path.exists(progress_path):
        with open(progress_path, "r", encoding="utf-8") as f:
            progress_data = json.load(f)
            namings_last_verse = progress_data.get("namings_last_verse", 0)
            collocations_last_verse = progress_data.get("collocations_last_verse", 0)
            categorization_last_verse = progress_data.get("categorization_last_verse", 0)

    # Fehlende Dateien anlegen
    initialize_files(paths)

    return (
        book_name,
        namings_last_verse,
        collocations_last_verse,
        categorization_last_verse,
        paths
    )

def initialize_files(paths):
    """Creates the project-specific JSON files if they do not already exist."""

    def create_if_missing(path, content):
        if not os.path.exists(path):
            with open(path, "w", encoding="utf-8") as f:
                json.dump(content, f, indent=4, ensure_ascii=False)

    create_if_missing(paths["progress_json"], {
        "namings_last_verse": 0,
        "collocations_last_verse": 0,
        "categorization_last_verse": 0
    })

    create_if_missing(paths["missing_namings_json"], [])
    create_if_missing(paths["collocations_json"], [])
    create_if_missing(paths["categorization_json"], [])

def load_data(load_excel: bool = False, load_tei: bool = False) -> DataType:
    """Interactively loads Excel and TEI files if requested by the configuration."""
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    data: DataType = {"excel": None, "excel_path": None, "xml": None}

    # 1. Load or create Excel file
    if load_excel:
        excel_path = filedialog.askopenfilename(
            title="Select the Excel file with naming data",
            initialdir=os.getcwd(),
            filetypes=[("Excel files", "*.xlsx")]
        )

        if excel_path:
            while True:
                try:
                    df = pd.read_excel(excel_path)
                    df = check_required_columns(df)
                    data["excel"] = df
                    data["excel_path"] = excel_path
                    print(f"✅ Excel file loaded: {os.path.basename(excel_path)}")
                    break
                except PermissionError:
                    print("❌ The Excel file is currently open or locked. Please close it and try again.")
                    retry = input("🔁 Retry file selection? (y/n): ").strip().lower()
                    if retry == "y":
                        excel_path = filedialog.askopenfilename(
                            title="Re-select the Excel file",
                            initialdir=os.getcwd(),
                            filetypes=[("Excel files", "*.xlsx")]
                        )
                        if not excel_path:
                            print("⚠️ No file selected – aborting.")
                            break
                    else:
                        break
                except Exception as e:
                    print(f"❌ Error loading Excel file: {e}")
                    break
        else:
            print("⚠️ No Excel file selected.")
            create_new = input("Would you like to create a new Excel file instead? (y/n): ").strip().lower()
            if create_new == "y":
                save_path = filedialog.asksaveasfilename(
                    title="Choose save location for the new Excel file",
                    defaultextension=".xlsx",
                    initialdir=os.getcwd(),
                    filetypes=[("Excel files", "*.xlsx")]
                )
                if save_path:
                    try:
                        template_path = os.path.join(os.getcwd(), "template_excel.xlsx")
                        wb = load_workbook(template_path)
                        wb.save(save_path)
                        df = pd.read_excel(save_path)
                        df = check_required_columns(df)
                        data["excel"] = df
                        data["excel_path"] = save_path
                        print(f"✅ New Excel file created: {os.path.basename(save_path)}")
                    except Exception as e:
                        print(f"❌ Error while creating the new file: {e}")
                else:
                    print("⚠️ No save location selected.")

    # 2. Load TEI-XML file
    if load_tei:
        xml_path = filedialog.askopenfilename(
            title="Select the TEI-XML file",
            initialdir=os.getcwd(),
            filetypes=[("XML files", "*.xml")]
        )
        if xml_path:
            try:
                tree = ET.parse(xml_path)
                root_elem = tree.getroot()
                root_elem = normalize_tei_text(root_elem)
                data["xml"] = root_elem
                print(f"✅ XML file loaded: {os.path.basename(xml_path)}")
                data["tei_path"] = xml_path

            except Exception as e:
                print(f"❌ Error loading XML file: {e}")
        else:
            print("⚠️ No XML file selected.")

    return data

def sorted_entries(entries: list) -> list:
    """
    Returns a sorted copy of the entries – by verse number and naming value.
    This enables stable comparison between two lists.
    """
    return sorted(
        deepcopy(entries),
        key=lambda x: (
            x.get("Vers", 0),
            x.get("Eigennennung") or x.get("Bezeichnung") or x.get("Erzähler") or ""
        )
    )

def check_required_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Checks whether all required columns are present in the DataFrame.
    If columns are missing, the user is prompted whether to add them automatically.
    Returns the (potentially extended) DataFrame.
    """
    required_columns = [
        "benannte figur",
        "vers",
        "eigennennung",
        "nennende figur",
        "bezeichnung",
        "erzähler",
        "kollokationen"
    ]

    current_columns_lower = [col.lower() for col in df.columns]
    missing_columns = [col for col in required_columns if col not in current_columns_lower]

    if not missing_columns:
        print("✅ All required columns are present.")
        return df

    print("⚠️ The following required columns are missing:")
    for col in missing_columns:
        print(f"   – {col}")

    for col in missing_columns:
        answer = input(f"Do you want to add the column \"{col}\" automatically? (y/n): ").strip().lower()
        if answer == "y":
            df[col] = ""
            print(f"➕ Column \"{col}\" added (empty).")
        else:
            print(f"⚠️ Column \"{col}\" remains missing.")

    return df

def normalize_text(text):
    """Normalizes a given text according to predefined rules."""
    substitutions = {
        'æ': 'ae', 'œ': 'oe',
        'é': 'e', 'è': 'e', 'ë': 'e', 'á': 'a', 'à': 'a',
        'û': 'u', 'î': 'i', 'â': 'a', 'ô': 'o', 'ê': 'e',
        'ü': 'u', 'ö': 'o', 'ä': 'a',
        'ß': 'ss',
        'iu': 'ie', 'üe': 'ue'
    }

    if not text:
        return ""

    text = text.lower()
    for old, new in substitutions.items():
        text = text.replace(old, new)

    text = re.sub(r'\bv\b', 'f', text)  # Replace 'v' at the beginning of words with 'f'
    text = re.sub(r'\s+', ' ', text)    # Collapse multiple spaces

    return text

def get_first_valid_text(*fields):
    """Returns the first field that is a non-empty string, skipping over NaN etc."""
    for f in fields:
        if isinstance(f, str) and f.strip():
            return f
    return ""

def normalize_tei_text(root):
    """Normalizes all text within the TEI file."""
    if root is None:
        return None

    normalized_lines = []
    for seg in root.findall('.//tei:seg', tei_ns):
        if seg.text:
            normalized_text = normalize_text(seg.text)
            seg.text = normalized_text
            normalized_lines.append(normalized_text)

    print("✅ TEI text has been normalized.")

    return root

def save_progress(
    missing_namings,
    last_processed_verse,
    paths,
    previous_verse=None,
    previous_namings=None,
    collocation_data=None,
    previous_collocations=None,
    categorized_entries=None,
    previous_categorized_entries=None,
    check_namings=False,
    perform_collocations=False,
    perform_categorization=False
):

    """
    Saves progress, namings, and optionally collocations or categorizations,
    only if there are changes compared to the previous state.
    """

    # Load existing progress file (if available)
    progress_data = {}
    if os.path.exists(paths["progress_json"]):
        with open(paths["progress_json"], "r", encoding="utf-8") as f:
            progress_data = json.load(f)

    # Update the respective last-verse value only if it changed
    if previous_verse is None or last_processed_verse != previous_verse:
        if check_namings:
            progress_data["namings_last_verse"] = last_processed_verse
        if perform_collocations:
            progress_data["collocations_last_verse"] = last_processed_verse
        if perform_categorization:
            progress_data["categorization_last_verse"] = last_processed_verse

        with open(paths["progress_json"], "w", encoding="utf-8") as f:
            json.dump(progress_data, f, indent=4, ensure_ascii=False)

    # 📌 Save namings – only if changed
    if previous_namings is None or sorted_entries(missing_namings) != sorted_entries(previous_namings):
        with open(paths["missing_namings_json"], "w", encoding="utf-8") as f:
            json.dump(missing_namings, f, indent=4, ensure_ascii=False)

    # 📌 Save collocations – only if provided and changed
    if collocation_data is not None:
        if previous_collocations is None or collocation_data != previous_collocations:
            with open(paths["collocations_json"], "w", encoding="utf-8") as f:
                json.dump(collocation_data, f, indent=4, ensure_ascii=False)

    # 📌 Save categorization – only if provided and changed
    if categorized_entries is not None:
        if previous_categorized_entries is None or sorted_entries(categorized_entries) != sorted_entries(previous_categorized_entries):
            with open(paths["categorization_json"], "w", encoding="utf-8") as f:
                json.dump(categorized_entries, f, indent=4, ensure_ascii=False)

def load_or_extend_naming_dict():
    """
    Loads or creates a central dictionary with character namings from Excel files.
    Returns a dict with structure {'Included Books': [...], 'Namings': {book: [namings, ...]}}
    """
    os.makedirs("data", exist_ok=True)
    dict_path = os.path.join("data", "naming_dict.json")

    # Load existing dict or create new one
    if os.path.exists(dict_path):
        with open(dict_path, "r", encoding="utf-8") as f:
            naming_dict = json.load(f)
        print(f"📚 A naming dictionary was found.")
        book_list = naming_dict.get("Included Books", [])
        if book_list:
            print(f"👉 Included books: {', '.join(book_list)}")
        else:
            print("👉 Included books: [empty]")
        extend = input("Do you want to add a file? (y/n): ").strip().lower()
    else:
        naming_dict = {"Included Books": [], "Namings": {}}
        print("❗ No naming dictionary found.")
        extend = "y"

    while extend == "y":
        print("📂 Please select an Excel file with naming data.")
        tk.Tk().withdraw()
        file_path = filedialog.askopenfilename(title="Select Excel file", filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            print("⚠️ No file selected. Operation cancelled.")
            break

        book_name = input("What is the name of the book? (e.g., Eneasroman): ").strip()

        namings = []

        try:
            df = pd.read_excel(file_path)
            relevant_columns = ["Eigennennung", "Bezeichnung", "Erzähler"]
            namings = []

            for column in relevant_columns:
                if column in df.columns:
                    namings.extend(df[column].dropna().tolist())

            # Remove duplicates and normalize
            namings = list(set(str(f).strip().lower() for f in namings if str(f).strip()))


        except PermissionError:
            print("❌ The Excel file is currently open or locked.")
            print("🔁 Please close the file and select it again.")
            file_path = filedialog.askopenfilename(
                title="Re-select the Excel file with namings",
                initialdir=os.getcwd(),
                filetypes=[("Excel files", "*.xlsx")]
            )
            if not file_path:
                print("⚠️ No file selected – aborting.")
                break

        except Exception as e:
            print(f"❌ Error while reading the file: {e}")
            break

        naming_dict["Included Books"].append(book_name)
        naming_dict["Namings"][book_name] = namings
        print(f"✅ Book '{book_name}' added with {len(namings)} namings.")

        extend = input("Do you want to add another file? (y/n): ").strip().lower()

    with open(dict_path, "w", encoding="utf-8") as f:
        json.dump(naming_dict, f, indent=4, ensure_ascii=False)
        print(f"💾 Current dictionary saved at: {dict_path}")

    return naming_dict

def ask_config_interactively(config_path):
    """
    Collects or reuses configuration and loads Excel/TEI files as configured.
    Returns a tuple: (config_data, data_dict)
    """
    config_data = {}
    data: DataType = {"excel": None, "excel_path": None, "xml": None}

    if os.path.exists(config_path):
        reuse = input("⚙️ A configuration for this book was found. Do you want to reuse the previous settings? (y/n): ").strip().lower()
        if reuse == "y":
            with open(config_path, "r", encoding="utf-8") as f:
                config_data = json.load(f)

            # 🟩 Excel automatisch laden, wenn Pfad bekannt & gültig
            excel_path = config_data.get("excel_path")
            if config_data.get("load_excel") and excel_path and os.path.exists(excel_path):
                try:
                    df = pd.read_excel(excel_path)
                    df = check_required_columns(df)
                    data["excel"] = df
                    data["excel_path"] = excel_path
                    print(f"✅ Excel file reloaded: {os.path.basename(excel_path)}")
                except PermissionError:
                    print(f"❌ Excel file is currently open or locked: {excel_path}")
                    retry = input("🔁 Retry file selection? (y/n): ").strip().lower()
                    if retry == "y":
                        partial = load_data(load_excel=True, load_tei=False)
                        if partial.get("excel") is not None:
                            data["excel"] = partial["excel"]
                            data["excel_path"] = partial.get("excel_path")
                            config_data["excel_path"] = partial.get("excel_path")
                except Exception as e:
                    print(f"❌ Failed to reload Excel: {e}")
            elif config_data.get("load_excel"):
                print(f"⚠️ Excel file not found at saved path: {excel_path}")

            # 🟩 TEI automatisch laden
            tei_path = config_data.get("tei_path")
            if config_data.get("load_tei") and tei_path and os.path.exists(tei_path):
                try:
                    tree = ET.parse(tei_path)
                    root_elem = tree.getroot()
                    root_elem = normalize_tei_text(root_elem)
                    data["xml"] = root_elem
                    print(f"✅ TEI file reloaded: {os.path.basename(tei_path)}")
                except Exception as e:
                    print(f"❌ Failed to reload TEI: {e}")
            elif config_data.get("load_tei"):
                print(f"⚠️ TEI file not found at saved path: {tei_path}")

            return config_data, data
        else:
            print("🛠 Reusing declined – please define new settings.")
    else:
        print("🛠 No existing config found – please define new settings.")

    # 🔹 Excel laden (nach Eingabe)
    config_data["load_excel"] = input(
        "Do you want to load an Excel file with existing naming data? (y/n): ").strip().lower() == "y"
    if config_data["load_excel"]:
        partial = load_data(load_excel=True, load_tei=False)

        if partial.get("excel") is not None:
            data["excel"] = partial["excel"]
            data["excel_path"] = partial.get("excel_path")
            config_data["excel_path"] = partial.get("excel_path")
        else:
            print("❌ No Excel file was loaded. Disabling Excel-related processing.")
            config_data["load_excel"] = False

    # 🔹 TEI laden (nach Eingabe)
    config_data["load_tei"] = input("Do you want to load the corresponding TEI file? (y/n): ").strip().lower() == "y"
    if config_data["load_tei"]:
        partial = load_data(load_excel=False, load_tei=True)

        if partial.get("xml") is not None:
            data["xml"] = partial["xml"]
            data["tei_path"] = partial.get("tei_path") or partial.get("xml_path")
            config_data["tei_path"] = data["tei_path"]
        else:
            print("❌ No TEI file was loaded. Disabling TEI-related processing.")
            config_data["load_tei"] = False

    config_data["check_namings"] = input("Should namings be checked and added? (y/n): ").strip().lower() == "y"
    config_data["fill_collocations"] = input("Should empty collocations be filled? (y/n): ").strip().lower() == "y"
    config_data["do_categorization"] = input("Should namings be lemmatized and categorized? (y/n): ").strip().lower() == "y"

    save_config(config_path, config_data)

    return config_data, data

def save_config(path, config_data):
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(config_data, f, indent=4, ensure_ascii=False)
        print(f"💾 Settings saved to: {path}")
    except Exception as e:
        print(f"❌ Failed to save config: {e}")

def search_tei_with_dict(
    df,
    root,
    naming_dict,
    last_verse,
    paths,
    missing_namings,
    collocation_data,
    check_namings=True,
    perform_collocations=False,
    perform_categorization=False,
    lemma_normalization=None,
    ignored_lemmas=None,
    lemma_categories=None,
    categorized_entries=None):

    """
    Iterates through the TEI text starting at the stored verse and performs the selected checks.
    """

    verse = root.findall('.//tei:l', tei_ns)
    if not verse:
        print("⚠️ No verses found.")
        return missing_namings

    start_index = next(
        (i for i, line in enumerate(verse) if int(line.get("n")) > last_verse),
        0
    )

    print(f"🔁 Starting iteration from verse {int(verse[start_index].get('n'))} (Index {start_index})")

    for line in verse[start_index:]:
        verse_number = int(line.get("n"))

        verse_text = ' '.join([seg.text for seg in line.findall(".//tei:seg", tei_ns) if seg.text])
        normalized_verse = normalize_text(verse_text)

        if check_namings:
            missing_namings = check_and_extend_namings(
                verse_number, verse_text, normalized_verse, df, naming_dict, missing_namings, root, paths
            )

        if perform_collocations:
            check_and_add_collocations(
                verse_number, df, collocation_data, root, paths
            )
        if perform_categorization:

            df_verse = df[df["Vers"] == verse_number]
            entries = df_verse.to_dict(orient="records")

            for entry in entries:
                source_text = normalize_text(get_first_valid_text(
                    entry.get("Erzähler"),
                    entry.get("Bezeichnung"),
                    entry.get("Eigennennung")
                ))
                if not source_text:
                    continue

                skip = False
                for e in categorized_entries:
                    if int(e.get("Vers", -1)) != verse_number:
                        continue

                    target_text = normalize_text(get_first_valid_text(
                        e.get("Erzähler"),
                        e.get("Bezeichnung"),
                        e.get("Eigennennung")
                    ))

                    if source_text == target_text and normalize_text(e.get("Benannte Figur", "")) == normalize_text(
                            entry.get("Benannte Figur", "")):
                        if any(
                                str(e.get(k, "")).strip()
                                for k in e.keys()
                                if k.startswith("Bezeichnung") or k.startswith("Epitheta")
                        ):
                            skip = True
                            break

                if skip:
                    continue

                annotated = lemmatize_and_categorize_entry(
                    entry, lemma_normalization, paths, ignored_lemmas, lemma_categories
                )
                if annotated:
                    categorized_entries.append(annotated)

        # Fortschritt speichern
        save_progress(
            missing_namings=missing_namings,
            last_processed_verse=verse_number,
            paths=paths,
            check_namings=check_namings,
            perform_collocations=perform_collocations,
            perform_categorization=perform_categorization
        )

    return missing_namings


def check_and_extend_namings(
    verse_number: int,
    verse_text: str,
    normalized_verse: str,
    df: pd.DataFrame,
    naming_dict: dict,
    missing_namings: list,
    root: Element,
    paths: dict
) -> list:
    """
    Checks whether a naming from the global dict appears in the current verse,
    but is not yet listed in Excel or in confirmed/rejected namings.
    If found: interactive confirmation and storage.
    """

    # 1. Extract namings from Excel for the current verse
    existing_namings = set()
    if "Vers" in df.columns:
        df_verse = df[df["Vers"] == verse_number]
        for column in ["Eigennennung", "Bezeichnung", "Erzähler"]:
            if column in df_verse.columns:
                values = df_verse[column].dropna().tolist()
                existing_namings.update(
                    normalize_text(str(value).strip()) for value in values if str(value).strip()
                )

    # 2. Extract and normalize namings from dict
    dict_namings = set()
    for book_list in naming_dict.get("Namings", {}).values():
        dict_namings.update(
            normalize_text(name.strip()) for name in book_list if name.strip()
        )

    # 3. Match check & user interaction
    for naming in dict_namings:
        if not naming:
            continue

        # skip if already handled in Excel (auch als Token-Menge)
        naming_tokens = set(naming.split())

        skip_existing = False
        for entry in existing_namings:
            entry_tokens = set(entry.split())
            if naming in entry or entry in naming:
                skip_existing = True
                break
            if naming_tokens <= entry_tokens or entry_tokens <= naming_tokens:
                skip_existing = True
                break

        if skip_existing:
            continue

        # skip if already handled in JSON
        skip = False
        for entry in missing_namings:
            if entry.get("Vers") == verse_number:
                values = [
                    entry.get("Eigennennung", ""),
                    entry.get("Bezeichnung", ""),
                    entry.get("Erzähler", "")
                ]
                if normalize_text(naming) in map(normalize_text, values):
                    skip = True
                    break
        if skip:
            continue

        if not re.search(rf'\b{re.escape(naming)}\b', normalized_verse):
            continue

        print("\n" + "-" * 60)
        print(f"❗ New naming found that is not listed in the Excel file!")
        print(f"🔍 Detected naming: \"{naming}\"")

        # 📖 Show context
        prev_line = root.find(f'.//tei:l[@n="{verse_number - 1}"]', tei_ns)
        if prev_line is not None:
            prev_text = ' '.join([seg.text for seg in prev_line.findall('.//tei:seg', tei_ns) if seg.text])
            print(f"📖 Previous verse ({verse_number - 1}): {prev_text}")

        highlighted = verse_text.replace(naming, f"\033[1m\033[93m{naming}\033[0m")
        print(f"📖 Verse ({verse_number}): {highlighted}")

        next_line = root.find(f'.//tei:l[@n="{verse_number + 1}"]', tei_ns)
        if next_line is not None:
            next_text = ' '.join([seg.text for seg in next_line.findall('.//tei:seg', tei_ns) if seg.text])
            print(f"📖 Next verse ({verse_number + 1}): {next_text}")

        # 🧍 Confirm with user
        confirm = input("Is this a missing naming? (y/n): ").strip().lower()
        if confirm == "n":
            missing_namings.append({
                "Vers": verse_number,
                "Eigennennung": naming,
                "Nennende Figur": "",
                "Bezeichnung": "",
                "Erzähler": "",
                "Status": "rejected"
            })
            save_progress(missing_namings, verse_number, paths)
            print("✅ Rejection saved.")
            continue

        extend = input("💡 Might this be a multi-word naming? Extend it? (y/n): ").strip().lower()
        if extend == "y":
            naming = input("✍ Enter the full naming: ").strip()

        print("Please choose the correct category:")
        print("[1] Eigennennung")
        print("[2] Bezeichnung")
        print("[3] Erzähler")
        print("[4] Skip")

        choice = input("👉 Your selection: ").strip()
        if choice == "4":
            continue

        named_entity = input("Enter the \"Benannte Figur\": ").strip()
        naming_entity = ""
        if choice == "2":
            naming_entity = input("Enter the \"Nennende Figur\": ").strip()

        entry = {
            "Benannte Figur": named_entity,
            "Vers": verse_number,
            "Eigennennung": naming if choice == "1" else "",
            "Nennende Figur": naming_entity,
            "Bezeichnung": naming if choice == "2" else "",
            "Erzähler": naming if choice == "3" else "",
            "Status": "confirmed"
        }

        # 📌 Optional: add collocation
        wants_collocation = input("📌 Do you want to add a collocation (context lines)? (y/n): ").strip().lower()
        if wants_collocation == "y":
            print("\n📖 Extended context (1–13):")
            context_lines = {}
            number = 1

            for i in range(6, 0, -1):
                line = root.find(f'.//tei:l[@n="{verse_number - i}"]', tei_ns)
                if line is not None:
                    text = ' '.join([seg.text for seg in line.findall('.//tei:seg', tei_ns) if seg.text])
                    context_lines[number] = text
                    print(f"[{number}] {text}")
                    number += 1

            context_lines[number] = verse_text
            print(f"[{number}] {verse_text}")
            number += 1

            for i in range(1, 7):
                line = root.find(f'.//tei:l[@n="{verse_number + i}"]', tei_ns)
                if line is not None:
                    text = ' '.join([seg.text for seg in line.findall('.//tei:seg', tei_ns) if seg.text])
                    context_lines[number] = text
                    print(f"[{number}] {text}")
                    number += 1

            selection = input("\n👉 Please enter the line number(s) (e.g., '5-7' or '6'): ").strip()
            selected = []

            try:
                if "-" in selection:
                    start, end = map(int, selection.split("-"))
                    selected = [context_lines[i] for i in range(start, end + 1) if i in context_lines]
                else:
                    idx = int(selection)
                    selected = [context_lines[idx]]
            except (ValueError, KeyError):
                print("⚠️ Invalid input – no collocation saved.")

            if selected:
                entry["Kollokation"] = ' / '.join(selected)

        missing_namings.append(entry)
        save_progress(missing_namings, verse_number, paths)
        print("✅ Entry saved.")

    return missing_namings

def load_missing_namings(path: str) -> list:
    """
    Loads missing or confirmed namings from a JSON file.
    Returns an empty list if the file doesn't exist or is invalid.
    """
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, FileNotFoundError):
            print("⚠️ Error while loading the JSON file – using an empty list.")
            return []
    else:
        return []

def get_verse_context(verse_number, root_tei):
    """Retrieves the surrounding 6 verses from the TEI file, numbered 1–13."""
    context = []
    verse_list = []

    for i in range(-6, 7):
        verse_id = str(verse_number + i)  # must be string!
        line = root_tei.find(f'.//tei:l[@n="{verse_id}"]', tei_ns)

        if line is not None:
            text = normalize_text(' '.join([
                seg.text for seg in line.findall('.//tei:seg', tei_ns) if seg.text
            ]))
            verse_list.append(text)

    for i, verse in enumerate(verse_list, start=1):
        context.append((i, verse))

    return context

def clean_cell_value(value):
    if pd.isna(value) or value is None:
        return ""
    return normalize_text(str(value).strip())

def check_and_add_collocations(verse_number, df, collocation_data, root, paths):
    """Checks whether a collocation should be added – if so, prompts for user input."""

    rows = df[df["Vers"] == verse_number]
    if rows.empty:
        return None

    # Check if already handled via Excel
    row = rows.iloc[0]
    if pd.notna(row.get("Kollokationen")) and str(row["Kollokationen"]).strip() != "":
        return None

    # Check if already handled via JSON
    if any(
            int(entry.get("Vers", -1)) == verse_number and str(entry.get("Kollokationen", "")).strip()
            for entry in collocation_data
    ):
        return None

    if any(
            str(entry.get("Vers")) == str(verse_number) and str(entry.get("Kollokationen", "")).strip()
            for entry in collocation_data
    ):
        return None  # already handled

    naming = clean_cell_value(row.get("Eigennennung")) \
             or clean_cell_value(row.get("Bezeichnung")) \
             or clean_cell_value(row.get("Erzähler"))

    named_entity = clean_cell_value(row.get("Benannte Figur"))

    context = get_verse_context(verse_number, root)

    collocations = ask_for_collocations(verse_number, named_entity, naming, context)

    collocation_data.append({
        "Vers": verse_number,
        "Kollokationen": collocations
    })

    # 📝 Immediately save progress
    with open(paths["collocations_json"], "w", encoding="utf-8") as f:
        json.dump(collocation_data, f, indent=4, ensure_ascii=False)

    return True

def ask_for_collocations(verse_number, named_entity, naming, context):
    """Displays the context around a verse and interactively asks for relevant collocations."""

    print(f"\n🟡 Empty collocation field detected in verse {verse_number}!")
    if named_entity or naming:
        print(f"👤 {named_entity}: {naming}\n")

    for number, text in context:
        if naming:
            # Highlight naming
            highlighted = text.replace(str(naming), f"\033[1;33m{naming}\033[0m")
        else:
            highlighted = text
        print(f"{number}. {highlighted}")

    user_input = input("\n👉 Please enter the number(s) of the relevant lines (e.g., '5' or '5-7'): ")

    selected = []

    try:
        if "-" in user_input:
            start, end = map(int, user_input.split("-"))
            selected = [text for number, text in context if start <= number <= end]
        else:
            number = int(user_input)
            selected = [text for num, text in context if num == number]
    except (ValueError, StopIteration):
        print("⚠️ Invalid input. Please enter a single number or a range.")
        return ""

    return " / ".join(selected)

def load_collocations_json(file_path):
    """Loads existing collocations from a JSON file – or returns an empty list if unavailable."""
    if os.path.exists(file_path):
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []

def load_ignored_lemmas(path="ignored_lemmas.json"):
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
            return set(data) if isinstance(data, list) else set(data.keys())
    return set()

def load_lemma_categories(path="lemma_categories.json"):
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def load_json_annotations(path):
    if os.path.exists(path):
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []

def save_json_annotations(path, annotations):
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(annotations, f, ensure_ascii=False, indent=2)

def lemmatize_and_categorize_entry(entry, lemma_normalization, paths, ignored_lemmas=None, lemma_categories=None):
    if lemma_normalization is None:
        lemma_normalization = load_lemma_normalization(paths["lemma_normalization_json"])

    if ignored_lemmas is None:
        ignored_lemmas = load_ignored_lemmas(paths["ignored_lemmas_json"])

    if lemma_categories is None:
        lemma_categories = load_lemma_categories(paths["lemma_categories_json"])

    text = get_first_valid_text(
        entry.get("Erzähler"),
        entry.get("Bezeichnung"),
        entry.get("Eigennennung")
    )

    if not text:
        print("⚠ No text to annotate – entry skipped.\n")
        return None

    print("\n" + "=" * 60)
    print(f"▶ Verse: {entry.get('Vers')}")
    print(f"▶ Named Entity: {entry.get('Benannte Figur')}")

    first_text = get_first_valid_text(
        entry.get("Eigennennung"),
        entry.get("Bezeichnung"),
        entry.get("Erzähler")
    )

    typ = "(unbestimmt)"

    if first_text == entry.get("Eigennennung"):
        typ = "Eigennennung"
    elif first_text == entry.get("Bezeichnung"):
        typ = "Bezeichnung"
    elif first_text == entry.get("Erzähler"):
        typ = "Erzähler"

    print(f"▶ Type: {typ}")

    print(f"\n▶ Original text: {text}")

    tokens = [t for t in tokenize(text.lower()) if t.isalpha()]

    # Filter only real word tokens
    missing = [t for t in tokens if t.isalpha() and t not in lemma_normalization]

    if missing:
        while True:
            print(f"\n▶ Please add lemma(ta) for {', '.join(missing)} (comma-separated):")
            user_input = input("> ").strip()
            new_lemmata = [l.strip() for l in user_input.split(",") if l.strip()]
            if len(new_lemmata) == len(missing):
                break
            print(
                f"⚠ Number of lemmata ({len(new_lemmata)}) doesn't match number of tokens ({len(missing)}). Please try again.")

        for token, lemma in zip(missing, new_lemmata):
            lemma_normalization[token] = lemma
        save_lemma_normalization(lemma_normalization, path=paths["lemma_normalization_json"])

    lemmata = [lemma_normalization.get(t, t) for t in tokens]
    print(f"\n▶ Lemma: {', '.join(lemmata)}\n")

    while True:
        designations, epithets = run_categorization(
            lemmata, lemma_categories, ignored_lemmas, paths
        )

        if not designations and not epithets:
            print("⚠ No entry – please review and confirm again.")
            confirm = input("Really skip this entry? [y = yes / n = no]: ").strip().lower()
            if confirm == "y":
                print("⏭ Entry skipped.\n")
                return None
            else:
                return run_categorization(lemmata, lemma_categories, ignored_lemmas, paths)
        else:
            break

    if lemma_normalization:
        save_lemma_normalization(lemma_normalization, path=paths["lemma_normalization_json"])

    if ignored_lemmas:
        save_ignored_lemmas(ignored_lemmas, path=paths["ignored_lemmas_json"])

    if lemma_categories:
        save_lemma_categories(lemma_categories, path=paths["lemma_categories_json"])

    annotated_entry = {
        **entry,
        "Bezeichnung 1": designations[0] if len(designations) > 0 else "",
        "Bezeichnung 2": designations[1] if len(designations) > 1 else "",
        "Bezeichnung 3": designations[2] if len(designations) > 2 else "",
        "Bezeichnung 4": designations[3] if len(designations) > 3 else "",
        "Epitheta 1": epithets[0] if len(epithets) > 0 else "",
        "Epitheta 2": epithets[1] if len(epithets) > 1 else "",
        "Epitheta 3": epithets[2] if len(epithets) > 2 else "",
        "Epitheta 4": epithets[3] if len(epithets) > 3 else "",
        "Epitheta 5": epithets[4] if len(epithets) > 4 else ""
    }

    # 💾 Kategorisierung direkt speichern
    existing = load_json_annotations(paths["categorization_json"])
    existing.append(annotated_entry)
    with open(paths["categorization_json"], "w", encoding="utf-8") as f:
        json.dump(existing, f, indent=4, ensure_ascii=False)

    print("✅ Entry saved.\n")
    return annotated_entry

def run_categorization(lemmata, lemma_categories, ignored_lemmas, paths):
    while True:
        designations = []
        epithets = []
        history = []
        i = 0

        while i < len(lemmata):
            lemma = lemmata[i]

            if lemma in ignored_lemmas:
                i += 1
                continue

            default = f"[{lemma_categories.get(lemma, '')}]" if lemma in lemma_categories else ""
            print(f"{lemma:<12} → {default} ", end="")
            user_input = input().strip()

            if user_input == "<":
                if i == 0 or not history:
                    print("↩️  Already at beginning – can't step back.")
                    continue
                i -= 1
                last_action = history.pop()
                if last_action["type"] == "a":
                    designations.pop()
                elif last_action["type"] == "e":
                    epithets.pop()
                elif last_action["type"] == "ignore":
                    ignored_lemmas.discard(last_action["lemma"])
                    save_ignored_lemmas(ignored_lemmas, path=paths["ignored_lemmas_json"])
                elif last_action["type"] == "override":
                    del lemma_categories[last_action["lemma"]]
                    save_lemma_categories(lemma_categories, path=paths["lemma_categories_json"])
                continue

            if user_input == "" and default:
                if default == "[a]":
                    designations.append(lemma)
                    history.append({"type": "a", "lemma": lemma})
                elif default == "[e]":
                    epithets.append(lemma)
                    history.append({"type": "e", "lemma": lemma})
                i += 1
                continue

            if user_input == "":
                confirm_ignore = input(f"⚠️ Really ignore lemma “{lemma}”? [y/n]: ").strip().lower()
                if confirm_ignore == "y":
                    ignored_lemmas.add(lemma)
                    save_ignored_lemmas(ignored_lemmas, path=paths["ignored_lemmas_json"])
                    print(f"ℹ️ Lemma “{lemma}” added to ignore list.")
                    history.append({"type": "ignore", "lemma": lemma})
                    i += 1
                    continue
                else:
                    print("↩️  Skipped ignoring – please choose a category or go back.\n")
                    continue

            if user_input in ("a", "e"):
                if user_input == "a":
                    designations.append(lemma)
                else:
                    epithets.append(lemma)
                lemma_categories[lemma] = user_input
                save_lemma_categories(lemma_categories, path=paths["lemma_categories_json"])
                history.append({"type": user_input, "lemma": lemma})
                i += 1
                continue

            correction = user_input
            cat = ""
            while cat not in ("a", "e"):
                cat = input(f'Define category for “{correction}” [a/e]: ').strip().lower()

            if cat == "a":
                designations.append(correction)
            else:
                epithets.append(correction)

            lemma_categories[correction] = cat
            save_lemma_categories(lemma_categories)
            history.append({"type": "override", "lemma": correction})
            i += 1

        return designations, epithets

def tokenize(text):
    return re.findall(r'\w+|[^\w\s]', text, re.UNICODE)

def load_lemma_normalization(path="lemma_normalization.json"):
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_lemma_normalization(data, path="lemma_normalization.json"):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def save_ignored_lemmas(data, path="ignored_lemmas.json"):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(sorted(data), f, ensure_ascii=False, indent=2)

def save_lemma_categories(data, path="lemma_categories.json"):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def export_all_data_to_new_excel(paths, options):
    """
    Integrates confirmed namings, adds collocations, and creates a lemmatized worksheet.

    :param paths: Dictionary containing file paths (original_excel, json_benennungen, json_kollokationen, json_kategorisierung)
    :param options: Dictionary with Boolean flags for: namings, collocations, categorization
    """

    # Generate target path with "_final" suffix
    original_name = os.path.basename(paths["original_excel"])
    target_name = original_name.replace(".xlsx", "_final.xlsx")

    # 🔧 Ensure output directory exists
    output_dir = "/mnt/endproduct"
    os.makedirs(output_dir, exist_ok=True)

    target_path = os.path.join(output_dir, str(target_name))

    # Copy Excel file
    shutil.copy(paths["original_excel"], target_path)

    wb = openpyxl.load_workbook(target_path)
    sheet = wb["Gesamt"]

    if options.get("benennungen", False):
        insert_namings(sheet, paths["json_benennungen"])

    if options.get("kollokationen", False):
        update_collocations(sheet, paths["json_kollokationen"])

    if options.get("kategorisierung", False):
        create_sheet_with_categorized_lemmata(wb, sheet, paths["json_kategorisierung"])

    wb.save(target_path)
    print(f"✅ Export completed: {target_path}")

def get_format_template(sheet, column_index):
    """
    Returns formatting template (font, alignment, border, number format) of the first filled cell in the column.
    """
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=column_index)
        if cell.value:
            if cell.has_style:
                return copy(cell.font), copy(cell.alignment), copy(cell.border), cell.number_format
    return None, None, None, None

def insert_namings(sheet, json_path):
    """
    Fügt bestätigte Benennungen in das Blatt 'Gesamt' ein.
    Formatierung pro Spalte wird über 'hole_formatvorlage()' übernommen.
    Neue Zeilen werden farblich hervorgehoben.
    """
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    confirmed_entries = [entry for entry in data if entry.get("Status") == "bestätigt"]
    if not confirmed_entries:
        print("ℹ️ Keine bestätigten Benennungen zum Einfügen.")
        return

    last_line = sheet.max_row + 1
    fill_color = PatternFill(start_color="4BACC6", end_color="4BACC6", fill_type="solid")

    for entry in confirmed_entries:
        new_line = [
            entry.get("Benannte Figur", ""),
            entry.get("Vers", ""),
            entry.get("Eigennennung", ""),
            entry.get("Nennende Figur", ""),
            entry.get("Bezeichnung", ""),
            entry.get("Erzähler", ""),
            entry.get("Kollokation", "")
        ]

        for col_num, value in enumerate(new_line, start=1):
            cell = sheet.cell(row=last_line, column=col_num, value=value)

            font_tpl, alignment_tpl, border_tpl, number_format_tpl = get_format_template(sheet, col_num)
            if font_tpl:
                cell.font = font_tpl
                cell.alignment = alignment_tpl
                cell.border = border_tpl
                cell.number_format = number_format_tpl

            cell.fill = fill_color

        last_line += 1

    print("✅ Benennungen erfolgreich ergänzt.")

def update_collocations(sheet, json_path):
    """
    Updates the 'Kollokationen' column using the JSON data.
    Formatting is copied from the first filled cell via get_format_template().
    """
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    header = [cell.value for cell in sheet[1]]
    try:
        verse_col = header.index("Vers") + 1
        collocation_col = header.index("Kollokationen") + 1
    except ValueError:
        print("❌ Columns 'Vers' or 'Kollokationen' not found!")
        return

    verse_to_rows = {}
    for row in range(2, sheet.max_row + 1):
        verse_value = sheet.cell(row=row, column=verse_col).value
        if verse_value is not None:
            verse_to_rows.setdefault(int(verse_value), []).append(row)

    font_tpl, alignment_tpl, border_tpl, number_format_tpl = get_format_template(sheet, collocation_col)

    updated_count = 0
    for entry in data:
        verse = entry["Vers"]
        new_value = entry["Kollokationen"]
        matching_rows = verse_to_rows.get(verse, [])
        for row in matching_rows:
            cell = sheet.cell(row=row, column=collocation_col, value=new_value)
            if font_tpl:
                cell.font = font_tpl
                cell.alignment = alignment_tpl
                cell.border = border_tpl
                cell.number_format = number_format_tpl
            updated_count += 1

    print(f"✅ {updated_count} collocations successfully updated.")

def create_sheet_with_categorized_lemmata(wb, original_sheet, json_path):
    """
    Creates a new worksheet 'lemmatized' with structured designations and epithets
    based on the data from the JSON file. Column formatting is applied using get_format_template().
    """
    with open(json_path, "r", encoding="utf-8") as f:
        annotations = json.load(f)

    # Reference original sheet
    ws_original = original_sheet

    # Delete existing 'lemmatized' sheet if present
    if 'lemmatisiert' in wb.sheetnames:
        del wb['lemmatisiert']

    # Create copy of original sheet
    ws_new = wb.copy_worksheet(ws_original)
    ws_new.title = 'lemmatisiert'

    # Position worksheet directly after 'Gesamt'
    if 'Gesamt' in wb.sheetnames:
        idx = wb.sheetnames.index('Gesamt') + 1
        wb._sheets.insert(idx, wb._sheets.pop(wb.sheetnames.index('lemmatisiert')))

    headers = [
        "Benannte Figur", "Vers", "Eigennennung", "Nennende Figur", "Bezeichnung", "Erzähler",
        "Bezeichnung 1", "Bezeichnung 2", "Bezeichnung 3", "Bezeichnung 4",
        "Epitheta 1", "Epitheta 2", "Epitheta 3", "Epitheta 4", "Epitheta 5"
    ]

    df_new = pd.DataFrame(annotations)
    for col in headers:
        if col not in df_new.columns:
            df_new[col] = ""

            if col == "Nennende Figur" and 'Gesamt' in wb.sheetnames:
                ws_gesamt = wb['Gesamt']
                header_row = [cell.value for cell in ws_gesamt[1]]
                verse_idx = header_row.index("Vers") + 1
                naming_idx = header_row.index("Nennende Figur") + 1

                verse_to_namer = {}
                for row in ws_gesamt.iter_rows(min_row=2):
                    verse_val = row[verse_idx - 1].value
                    naming_val = row[naming_idx - 1].value
                    if verse_val is not None:
                        verse_to_namer[verse_val] = naming_val

                df_new["Nennende Figur"] = df_new["Vers"].map(verse_to_namer).fillna("")

    df_new = df_new[headers]

    # Remove old 'Kollokationen' column if present
    for col in ws_new.iter_cols(min_row=1, max_row=1):
        if col[0].value == "Kollokationen":
            ws_new.delete_cols(col[0].column, 1)
            break

    # Remove duplicate columns
    existing_headers = [cell.value for cell in ws_new[1]]
    redundant_cols = [i + 1 for i, h in enumerate(existing_headers) if h in headers]
    for idx in sorted(redundant_cols, reverse=True):
        ws_new.delete_cols(idx)

    # Insert new headers and formatted data
    for col_offset, header in enumerate(headers):
        col_index = col_offset + 1
        ws_new.cell(row=1, column=col_index, value=header)
        font_tpl, alignment_tpl, border_tpl, number_format_tpl = get_format_template(ws_original, 1)

        for row_idx, value in enumerate(df_new[header], start=2):
            cell = ws_new.cell(row=row_idx, column=col_index, value=value)
            if font_tpl:
                cell.font = font_tpl
                cell.alignment = alignment_tpl
                cell.border = border_tpl
                cell.number_format = number_format_tpl

    ws_new.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    print("✅ Worksheet 'lemmatisiert' successfully created.")

def main():
    # 🔹 1. Initialization: book selection, paths, last verse
    book_name, namings_last_verse, collocations_last_verse, categorization_last_verse, paths = initialize_project()

    # 🔹 2. Load global naming dictionary (from all books)
    naming_dict = load_or_extend_naming_dict()

    # 🔹 3. Konfiguration abfragen (interaktiv oder reuse)
    config_data, data = ask_config_interactively(paths["config_json"])

    paths["original_excel"] = data.get("excel_path")
    df = data.get("excel")
    root = data.get("xml")

    # 🔹 Preload all JSON data for duplication checks (independent of mode)
    missing_namings = load_missing_namings(paths["missing_namings_json"])
    collocation_data = load_collocations_json(paths["collocations_json"])
    categorized_entries = load_json_annotations(paths["categorization_json"])

    # Defaults for optional tracking
    previous_namings = []
    previous_collocations = []
    previous_categorized_entries = []

    # Optional: lemma support for categorization
    lemma_normalization = None
    ignored_lemmas = None
    lemma_categories = None

    # 🔹 6. User-controlled analysis paths
    check_namings = config_data["check_namings"]
    fill_collocations = config_data["fill_collocations"]
    do_categorization = config_data["do_categorization"]

    # 🔹 7. Load previous data depending on paths
    if check_namings:
        previous_verse = namings_last_verse
        active_last_verse = namings_last_verse
        previous_namings = missing_namings.copy()

    elif fill_collocations:
        previous_verse = collocations_last_verse
        active_last_verse = collocations_last_verse
        previous_collocations = collocation_data.copy()

    elif do_categorization:
        previous_verse = categorization_last_verse
        active_last_verse = categorization_last_verse
        previous_categorized_entries = categorized_entries.copy()
        lemma_normalization = load_lemma_normalization(paths["lemma_normalization_json"])
        ignored_lemmas = load_ignored_lemmas(paths["ignored_lemmas_json"])
        lemma_categories = load_lemma_categories(paths["lemma_categories_json"])

    else:
        previous_verse = 0
        active_last_verse = 0

    # 🔹 8. Process TEI and run requested analysis steps
    missing_namings = search_tei_with_dict(
        df=df,
        root=root,
        naming_dict=naming_dict,
        last_verse=active_last_verse,
        paths=paths,
        missing_namings=missing_namings,
        collocation_data=collocation_data,
        check_namings=check_namings,
        perform_collocations=fill_collocations,
        perform_categorization=do_categorization,
        lemma_normalization=lemma_normalization if do_categorization else None,
        ignored_lemmas=ignored_lemmas if do_categorization else None,
        lemma_categories=lemma_categories if do_categorization else None,
        categorized_entries=categorized_entries if do_categorization else None
    )

    # 🔹 9. Final save
    save_progress(
        missing_namings=missing_namings,
        last_processed_verse=active_last_verse,
        paths=paths,
        previous_verse=previous_verse,
        previous_namings=previous_namings,
        collocation_data=collocation_data,
        previous_collocations=previous_collocations,
        categorized_entries=categorized_entries,
        previous_categorized_entries=previous_categorized_entries,
        check_namings=check_namings,
        perform_collocations=fill_collocations,
        perform_categorization=do_categorization
    )

    # 🔹 10. Optional export
    export = input("Do you want to export all results? (y/n): ").strip().lower() == "y"
    if export:
        paths["original_excel"] = data["excel_path"]
        options = {
            "benennungen": check_namings,
            "kollokationen": fill_collocations,
            "kategorisierung": do_categorization
        }
        export_all_data_to_new_excel(paths, options)


if __name__ == "__main__":
    main()
