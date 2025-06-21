# Built-in libraries
import os
import re
import csv
import json
import time
import math
import shutil
import difflib
import webbrowser
from copy import copy, deepcopy
from collections import Counter

# Type annotations
from typing import Union, List

# GUI and file dialogs
import tkinter as tk
from tkinter import filedialog

# Third-party libraries
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border

import plotly.express as px

# XML and TEI handling
import xml.etree.ElementTree as ET
from xml.etree.ElementTree import Element


DataType = dict[str, Union[pd.DataFrame, Element, str, None]]
tei_ns = {'tei': 'http://www.tei-c.org/ns/1.0'}

def load_cached_json_data(paths: dict) -> tuple[list, dict, dict]:
    """
    Loads cached JSON data for:
    - missing naming variants
    - collocation data
    - categorized entries

    Args:
        paths (dict): dictionary containing all relevant JSON paths

    Returns:
        tuple:
            missing_naming_variants (dict)
            collocation_data (dict)
            categorized_entries (dict)
    """
    missing_naming_variants = load_missing_naming_variants(paths["missing_naming_variants_json"])
    collocation_data = load_collocations_json(paths["collocations_json"])
    categorized_entries = load_json_annotations(paths["categorization_json"])

    return missing_naming_variants, collocation_data, categorized_entries

def export_if_enabled(
    data: dict,
    paths: dict,
    book_name: str,
    check_naming_variants: bool,
    perform_collocations: bool,
    perform_categorization: bool
):
    """
    Asks the user whether to export results and triggers export if confirmed.

    Args:
        data (dict): contains original Excel path info
        paths (dict): project paths for output
        book_name (str): name of the text being processed
        check_naming_variants (bool): whether naming variants were collected
        perform_collocations (bool): whether collocations were collected
        perform_categorization (bool): whether categorization was performed
    """
    export = ask_user_choice("Do you want to export all results? (y/n): ", ["y", "n"]) == "y"
    if not export:
        return

    options = {
        "benennungen": check_naming_variants,
        "kollokationen": perform_collocations,
        "kategorisierung": perform_categorization
    }

    paths["original_excel"] = data.get("excel_path")
    export_all_data_to_new_excel(book_name, paths, options)

def analyze_if_enabled(
    config_data: dict,
    paths: dict,
    data: dict,
    book_name: str
):
    """
    Asks the user whether to run an analysis after data collection and triggers it if confirmed.

    Args:
        config_data (dict): configuration settings
        paths (dict): all relevant file paths
        data (dict): loaded data (e.g. Excel, TEI)
        book_name (str): name of the current text
    """
    analyze_after = ask_user_choice("Do you want to run an analysis now? (y/n): ", ["y", "n"])
    if analyze_after == "y":
        run_analysis_menu(config_data, paths, data, book_name)

def main():
    # 🔹 1. Initialization: book selection, paths, last verse
    book_name, naming_variants_last_verse, collocations_last_verse, categorization_last_verse, paths = initialize_project()

    # 🔹 2. Load global naming dictionary (from all books)
    naming_variants_dict = load_or_extend_naming_variants_dict()

    # 🔹 3. Konfiguration abfragen (interaktiv oder reuse)
    config_data, data = ask_config_interactively(paths["config_json"])

    paths["original_excel"] = data.get("excel_path")
    df = data.get("excel")
    root = data.get("xml")

    missing_naming_variants, collocation_data, categorized_entries = load_cached_json_data(paths)

    # Defaults for optional tracking
    previous_naming_variants = []
    previous_collocations = []
    previous_categorized_entries = []

    # Optional: lemma support for categorization
    lemma_normalization = None
    ignored_lemmas = None
    lemma_categories = None

    mode = config_data.get("modus", "collect")

    if mode == "analyze":
        run_analysis_menu(config_data, paths, data, book_name)
        return  # skip rest of collection logic

    # 🔹 6. User-controlled analysis paths
    check_naming_variants = config_data["check_naming_variants"]
    fill_collocations = config_data["fill_collocations"]
    do_categorization = config_data["do_categorization"]

    # 🔹 7. Load previous data depending on paths
    if check_naming_variants:
        previous_verse = naming_variants_last_verse
        active_last_verse = naming_variants_last_verse
        previous_naming_variants = missing_naming_variants.copy()

    elif fill_collocations:
        previous_verse = collocations_last_verse
        active_last_verse = collocations_last_verse
        previous_collocations = collocation_data.copy()

    elif do_categorization:
        previous_verse = categorization_last_verse
        active_last_verse = categorization_last_verse
        previous_categorized_entries = categorized_entries.copy()
        lemma_normalization = load_lemma_normalization(paths["lemma_normalization_json"])
        ignored_lemmas = load_ignored_lemmas(paths["ignored_lemmas_json"])
        lemma_categories = load_lemma_categories(paths["lemma_categories_json"])

    else:
        previous_verse = 0
        active_last_verse = 0

    # 🔹 8. Process TEI and run requested analysis steps
    missing_naming_variants, collocation_data, categorized_entries = run_data_collection(
        df=df,
        root=root,
        naming_variants_dict=naming_variants_dict,
        last_verse=active_last_verse,
        paths=paths,
        missing_naming_variants=missing_naming_variants,
        collocation_data=collocation_data,
        check_naming_variants=check_naming_variants,
        perform_collocations=fill_collocations,
        perform_categorization=do_categorization,
        lemma_normalization=lemma_normalization if do_categorization else None,
        ignored_lemmas=ignored_lemmas if do_categorization else None,
        lemma_categories=lemma_categories if do_categorization else None,
        categorized_entries=categorized_entries if do_categorization else None
    )

    # 🔹 9. Final save
    save_progress(
        missing_naming_variants=missing_naming_variants,
        last_processed_verse=active_last_verse,
        paths=paths,
        previous_verse=previous_verse,
        previous_naming_variants=previous_naming_variants,
        collocation_data=collocation_data,
        previous_collocations=previous_collocations,
        categorized_entries=categorized_entries,
        previous_categorized_entries=previous_categorized_entries,
        check_naming_variants=check_naming_variants,
        perform_collocations=fill_collocations,
        perform_categorization=do_categorization
    )

    export_if_enabled(
        data,
        paths,
        book_name,
        check_naming_variants=check_naming_variants,
        perform_collocations=fill_collocations,
        perform_categorization=do_categorization
    )

    analyze_if_enabled(config_data, paths, data, book_name)


if __name__ == "__main__":
    main()
