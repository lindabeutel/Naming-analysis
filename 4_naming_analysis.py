import tkinter as tk
from tkinter import filedialog

import pandas as pd
import re
import json
import os
import shutil
import time
import csv
import difflib
import math

import xml.etree.ElementTree as ET
from xml.etree.ElementTree import Element

from copy import copy, deepcopy

from typing import Union, List

from collections import Counter

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles import Font, Alignment, Border
from openpyxl import load_workbook

DataType = dict[str, Union[pd.DataFrame, Element, str, None]]
tei_ns = {'tei': 'http://www.tei-c.org/ns/1.0'}

def get_valid_verse_number(value, fallback=-1):
    """
    Tries to parse a verse number as integer.
    Returns fallback if conversion fails (e.g. empty, NaN, wrong type).
    """
    try:
        return int(str(value).strip())
    except (ValueError, TypeError):
        return fallback

def safe_write_json(data, path, sort_keys=False, merge=False):
    for attempt in range(2):
        try:
            # Bestehende Daten einlesen, wenn merging aktiv ist
            if merge and os.path.exists(path):
                try:
                    with open(path, "r", encoding="utf-8") as f:
                        existing = json.load(f)
                except (FileNotFoundError, json.JSONDecodeError, PermissionError):
                    existing = [] if isinstance(data, (list, set)) else {}

                # Set → Liste umwandeln
                if isinstance(data, set):
                    data = list(data)

                # 🔄 Merge-Strategie
                if isinstance(data, list) and isinstance(existing, list):
                    if all(isinstance(x, dict) for x in data + existing):
                        # ✅ dedupliziere dicts nach (Vers, Benannte Figur)
                        seen = set()
                        merged = []
                        for entry in existing + data:
                            key = (
                                entry.get("Vers"),
                                entry.get("Benannte Figur"),
                                entry.get("Eigennennung") or entry.get("Bezeichnung") or entry.get("Erzähler")
                            )
                            if key not in seen:
                                merged.append(entry)
                                seen.add(key)
                        data = merged
                    else:
                        # list of str etc.
                        data = list(set(existing).union(set(data)))

                elif isinstance(data, dict) and isinstance(existing, dict):
                    existing.update(data)
                    data = existing

            elif isinstance(data, set):
                data = list(data)

            with open(path, "w", encoding="utf-8") as f:
                json.dump(
                    sorted(data) if sort_keys and isinstance(data, list) else data,
                    f,
                    ensure_ascii=False,
                    indent=2
                )
            return

        except PermissionError as e:
            if attempt == 0:
                print(f"⚠️ Zugriff verweigert auf {path}. Warte 1 Sekunde und versuche erneut...")
                time.sleep(1)
            else:
                print(f"❌ Zweiter Versuch fehlgeschlagen. Datei bleibt gesperrt: {path}")
                raise e

def safe_read_json(path, default=None):
    """
    Safely reads JSON from the given path. Returns `default` if file is missing or unreadable.
    """
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"⚠️ Datei nicht gefunden: {path} – es wird eine Ersatzstruktur verwendet.")
        return default if default is not None else {}
    except json.JSONDecodeError:
        print(f"⚠️ JSON-Fehler beim Lesen von {path} – leere Struktur wird verwendet.")
        return default if default is not None else {}
    except PermissionError:
        print(f"❌ Zugriff verweigert auf {path} – lese Vorgang abgebrochen.")
        return default if default is not None else {}

def ask_user_choice(prompt, valid_options):
    """
    Keeps asking for an entry until a valid option has been entered.
    """
    valid_options = [opt.lower() for opt in valid_options]
    while True:
        user_input = input(prompt).strip().lower()
        if user_input in valid_options:
            return user_input
        print(f"⚠️ Invalid input. Please select one of the following options: {', '.join(valid_options)}")

def initialize_project():
    book_name = input("Which book are we working on today? (e.g., Iwein): ").strip()
    book_name = book_name[0].upper() + book_name[1:]

    project_dir = os.path.join("data", book_name)
    os.makedirs(project_dir, exist_ok=True)
    os.makedirs("data", exist_ok=True)  # Für globale Dateien, falls nicht vorhanden

    config_path = os.path.join(project_dir, f"config_{book_name}.json")
    progress_path = os.path.join(project_dir, f"progress_{book_name}.json")

    paths = {
        # Projektbezogene Dateien
        "missing_naming_variants_json": os.path.join(project_dir, f"missing_naming_variants_{book_name}.json"),
        "progress_json": progress_path,
        "collocations_json": os.path.join(project_dir, f"collocations_{book_name}.json"),
        "categorization_json": os.path.join(project_dir, f"categorization_{book_name}.json"),
        "config_json": config_path,

        # Globale Dateien
        "lemma_normalization_json": os.path.join("data", "lemma_normalization.json"),
        "ignored_lemmas_json": os.path.join("data", "ignored_lemmas.json"),
        "lemma_categories_json": os.path.join("data", "lemma_categories.json")
    }

    # Fortschritt laden oder initialisieren
    namings_last_verse = 0
    collocations_last_verse = 0
    categorization_last_verse = 0

    if os.path.exists(progress_path):
        progress_data = safe_read_json(progress_path, default={})
        namings_last_verse = progress_data.get("namings_last_verse", 0)
        collocations_last_verse = progress_data.get("collocations_last_verse", 0)
        categorization_last_verse = progress_data.get("categorization_last_verse", 0)

    # Fehlende Dateien anlegen
    initialize_files(paths)

    return (
        book_name,
        namings_last_verse,
        collocations_last_verse,
        categorization_last_verse,
        paths
    )

def initialize_files(paths):
    """Creates the project-specific JSON files if they do not already exist."""

    def create_if_missing(path, content):
        if not os.path.exists(path):
            with open(path, "w", encoding="utf-8") as f:
                json.dump(content, f, indent=4, ensure_ascii=False)

    create_if_missing(paths["progress_json"], {
        "namings_last_verse": 0,
        "collocations_last_verse": 0,
        "categorization_last_verse": 0
    })

    create_if_missing(paths["missing_naming_variants_json"], [])
    create_if_missing(paths["collocations_json"], [])
    create_if_missing(paths["categorization_json"], [])

def load_data(load_excel: bool = False, load_tei: bool = False) -> DataType:
    """Interactively loads Excel and TEI files if requested by the configuration."""
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    data: DataType = {"excel": None, "excel_path": None, "xml": None}

    # 1. Load or create Excel file
    if load_excel:
        excel_path = filedialog.askopenfilename(
            title="Select the Excel file with naming data",
            initialdir=os.getcwd(),
            filetypes=[("Excel files", "*.xlsx")]
        )

        if excel_path:
            while True:
                try:
                    df = pd.read_excel(excel_path)
                    df = check_required_columns(df)
                    data["excel"] = df
                    data["excel_path"] = excel_path
                    print(f"✅ Excel file loaded: {os.path.basename(excel_path)}")
                    break
                except PermissionError:
                    print("❌ The Excel file is currently open or locked. Please close it and try again.")
                    retry = ask_user_choice("🔁 Retry file selection? (y/n): ", ["y", "n"])
                    if retry == "y":
                        excel_path = filedialog.askopenfilename(
                            title="Re-select the Excel file",
                            initialdir=os.getcwd(),
                            filetypes=[("Excel files", "*.xlsx")]
                        )
                        if not excel_path:
                            print("⚠️ No file selected – aborting.")
                            break
                    else:
                        break
                except Exception as e:
                    print(f"❌ Error loading Excel file: {e}")
                    break
        else:
            print("⚠️ No Excel file selected.")
            create_new = ask_user_choice("Would you like to create a new Excel file instead? (y/n): ", ["y", "n"])
            if create_new == "y":
                save_path = filedialog.asksaveasfilename(
                    title="Choose save location for the new Excel file",
                    defaultextension=".xlsx",
                    initialdir=os.getcwd(),
                    filetypes=[("Excel files", "*.xlsx")]
                )
                if save_path:
                    try:
                        template_path = os.path.join(os.getcwd(), "template_excel.xlsx")
                        wb = load_workbook(template_path)
                        wb.save(save_path)
                        df = pd.read_excel(save_path)
                        df = check_required_columns(df)
                        data["excel"] = df
                        data["excel_path"] = save_path
                        print(f"✅ New Excel file created: {os.path.basename(save_path)}")
                    except Exception as e:
                        print(f"❌ Error while creating the new file: {e}")
                else:
                    print("⚠️ No save location selected.")

    # 2. Load TEI-XML file
    if load_tei:
        xml_path = filedialog.askopenfilename(
            title="Select the TEI-XML file",
            initialdir=os.getcwd(),
            filetypes=[("XML files", "*.xml")]
        )
        if xml_path:
            try:
                tree = ET.parse(xml_path)
                root_elem = tree.getroot()
                root_elem = normalize_tei_text(root_elem)
                data["xml"] = root_elem
                print(f"✅ XML file loaded: {os.path.basename(xml_path)}")
                data["tei_path"] = xml_path

            except Exception as e:
                print(f"❌ Error loading XML file: {e}")
        else:
            print("⚠️ No XML file selected.")

    return data

def sorted_entries(entries: list) -> list:
    # Only keep entries with a valid numeric "Vers" value
    entries_clean = [
        e for e in deepcopy(entries)
        if isinstance(e, dict)
        and str(e.get("Vers", "")).strip().isdigit()
    ]

    return sorted(
        entries_clean,
        key=lambda x: (
            get_valid_verse_number(x.get("Vers")),
            get_first_valid_text(
                x.get("Eigennennung"),
                x.get("Bezeichnung"),
                x.get("Erzähler")
            ).strip().lower()
        )
    )


def check_required_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Checks whether all required columns are present in the DataFrame.
    If columns are missing, the user is prompted whether to add them automatically.
    Returns the (potentially extended) DataFrame.
    """
    required_columns = [
        "benannte figur",
        "vers",
        "eigennennung",
        "nennende figur",
        "bezeichnung",
        "erzähler",
        "kollokationen"
    ]

    current_columns_lower = [col.lower() for col in df.columns]
    missing_columns = [col for col in required_columns if col not in current_columns_lower]

    if not missing_columns:
        print("✅ All required columns are present.")
        return df

    print("⚠️ The following required columns are missing:")
    for col in missing_columns:
        print(f"   – {col}")

    for col in missing_columns:
        answer = input(f"Do you want to add the column \"{col}\" automatically? (y/n): ").strip().lower()
        if answer == "y":
            df[col] = ""
            print(f"➕ Column \"{col}\" added (empty).")
        else:
            print(f"⚠️ Column \"{col}\" remains missing.")

    return df

def normalize_text(text):
    """Normalizes a given text according to predefined rules."""
    substitutions = {
        'æ': 'ae', 'œ': 'oe',
        'é': 'e', 'è': 'e', 'ë': 'e', 'á': 'a', 'à': 'a',
        'û': 'u', 'î': 'i', 'â': 'a', 'ô': 'o', 'ê': 'e',
        'ü': 'u', 'ö': 'o', 'ä': 'a',
        'ß': 'ss',
        'iu': 'ie', 'üe': 'ue'
    }

    if not text:
        return ""

    text = text.lower()
    for old, new in substitutions.items():
        text = text.replace(old, new)

    text = re.sub(r'\bv\b', 'f', text)  # Replace 'v' at the beginning of words with 'f'
    text = re.sub(r'\s+', ' ', text)    # Collapse multiple spaces

    return text

def get_first_valid_text(*fields):
    """Returns the first field that is a non-empty string, skipping over NaN etc."""
    for f in fields:
        if isinstance(f, str) and f.strip():
            return f
    return ""

def normalize_tei_text(root):
    """Normalizes all text within the TEI file."""
    if root is None:
        return None

    normalized_lines = []
    for seg in root.findall('.//tei:seg', tei_ns):
        if seg.text:
            normalized_text = normalize_text(seg.text)
            seg.text = normalized_text
            normalized_lines.append(normalized_text)

    print("✅ TEI text has been normalized.")

    return root

def save_progress(
    missing_naming_variants,
    last_processed_verse,
    paths,
    previous_verse=None,
    previous_namings=None,
    collocation_data=None,
    previous_collocations=None,
    categorized_entries=None,
    previous_categorized_entries=None,
    check_namings=False,
    perform_collocations=False,
    perform_categorization=False
):

    """
    Saves progress, namings, and optionally collocations or categorizations,
    only if there are changes compared to the previous state.
    """

    # Load existing progress file (if available)
    progress_data = safe_read_json(paths["progress_json"], default={})

    # Update the respective last-verse value only if it changed
    if previous_verse is None or last_processed_verse != previous_verse:
        if check_namings:
            progress_data["namings_last_verse"] = last_processed_verse
        if perform_collocations:
            progress_data["collocations_last_verse"] = last_processed_verse
        if perform_categorization:
            progress_data["categorization_last_verse"] = last_processed_verse

        safe_write_json(progress_data, paths["progress_json"])

    if previous_namings is None or sorted_entries(missing_naming_variants) != sorted_entries(previous_namings):
        safe_write_json(missing_naming_variants, paths["missing_naming_variants_json"], merge=True)

    if collocation_data is not None:
        if previous_collocations is None or collocation_data != previous_collocations:
            safe_write_json(collocation_data, paths["collocations_json"], merge=True)

    if categorized_entries is not None:
        if previous_categorized_entries is None or sorted_entries(categorized_entries) != sorted_entries(
                previous_categorized_entries):
            safe_write_json(categorized_entries, paths["categorization_json"], merge=True)

def load_or_extend_naming_dict():
    """
    Loads or creates a central dictionary with character namings from Excel files.
    Returns a dict with structure {'Included Books': [...], 'Namings': {book: [namings, ...]}}
    """
    os.makedirs("data", exist_ok=True)
    dict_path = os.path.join("data", "naming_dict.json")

    # Load existing dict or create new one
    if os.path.exists(dict_path):
        naming_dict = safe_read_json(dict_path, default={"Included Books": [], "Namings": {}})
        print(f"📚 A naming dictionary was found.")
        book_list = naming_dict.get("Included Books", [])
        if book_list:
            print(f"👉 Included books: {', '.join(book_list)}")
        else:
            print("👉 Included books: [empty]")
        extend = ask_user_choice("Do you want to add a file? (y/n): ", ["y", "n"])
    else:
        naming_dict = {"Included Books": [], "Namings": {}}
        print("❗ No naming dictionary found.")
        extend = "y"

    while extend == "y":
        print("📂 Please select an Excel file with naming data.")
        tk.Tk().withdraw()
        file_path = filedialog.askopenfilename(title="Select Excel file", filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            print("⚠️ No file selected. Operation cancelled.")
            break

        book_name = input("What is the name of the book? (e.g., Eneasroman): ").strip()

        namings = []

        try:
            df = pd.read_excel(file_path)
            relevant_columns = ["Eigennennung", "Bezeichnung", "Erzähler"]
            namings = []

            for column in relevant_columns:
                if column in df.columns:
                    namings.extend(df[column].dropna().tolist())

            # Remove duplicates and normalize
            namings = list(set(str(f).strip().lower() for f in namings if str(f).strip()))


        except PermissionError:
            print("❌ The Excel file is currently open or locked.")
            print("🔁 Please close the file and select it again.")
            file_path = filedialog.askopenfilename(
                title="Re-select the Excel file with namings",
                initialdir=os.getcwd(),
                filetypes=[("Excel files", "*.xlsx")]
            )
            if not file_path:
                print("⚠️ No file selected – aborting.")
                break

        except Exception as e:
            print(f"❌ Error while reading the file: {e}")
            break

        naming_dict["Included Books"].append(book_name)
        naming_dict["Namings"][book_name] = namings
        print(f"✅ Book '{book_name}' added with {len(namings)} naming variants.")

        extend = ask_user_choice("Do you want to add another file? (y/n): ", ["y", "n"])

        safe_write_json(naming_dict, dict_path)
        print(f"💾 Current dictionary saved at: {dict_path}")

    return naming_dict

def ask_config_interactively(config_path):
    """
    Collects or reuses configuration and loads Excel/TEI files as configured.
    Returns a tuple: (config_data, data_dict)
    """
    config_data = {}
    data: DataType = {"excel": None, "excel_path": None, "xml": None}

    if os.path.exists(config_path):
        reuse = ask_user_choice("⚙️ A configuration for this book was found. Do you want to reuse the previous settings? (y/n): ", ["y", "n"])
        if reuse == "y":
            config_data = safe_read_json(config_path, default={})

            # 🟩 Excel automatisch laden, wenn Pfad bekannt & gültig
            excel_path = config_data.get("excel_path")
            if config_data.get("load_excel") and excel_path and os.path.exists(excel_path):
                try:
                    df = pd.read_excel(excel_path)
                    df = check_required_columns(df)
                    data["excel"] = df
                    data["excel_path"] = excel_path
                    print(f"✅ Excel file reloaded: {os.path.basename(excel_path)}")
                except PermissionError:
                    print(f"❌ Excel file is currently open or locked: {excel_path}")
                    retry = input("🔁 Retry file selection? (y/n): ").strip().lower()
                    if retry == "y":
                        partial = load_data(load_excel=True, load_tei=False)
                        if partial.get("excel") is not None:
                            data["excel"] = partial["excel"]
                            data["excel_path"] = partial.get("excel_path")
                            config_data["excel_path"] = partial.get("excel_path")
                except Exception as e:
                    print(f"❌ Failed to reload Excel: {e}")
            elif config_data.get("load_excel"):
                print(f"⚠️ Excel file not found at saved path: {excel_path}")

            # 🟩 TEI automatisch laden
            tei_path = config_data.get("tei_path")
            if config_data.get("load_tei") and tei_path and os.path.exists(tei_path):
                try:
                    tree = ET.parse(tei_path)
                    root_elem = tree.getroot()
                    root_elem = normalize_tei_text(root_elem)
                    data["xml"] = root_elem
                    print(f"✅ TEI file reloaded: {os.path.basename(tei_path)}")
                except Exception as e:
                    print(f"❌ Failed to reload TEI: {e}")
            elif config_data.get("load_tei"):
                print(f"⚠️ TEI file not found at saved path: {tei_path}")

            return config_data, data
        else:
            print("🛠 Reusing declined – please define new settings.")
    else:
        print("🛠 No existing config found – please define new settings.")

    # 🔹 Excel laden (nach Eingabe)
    config_data["load_excel"] = ask_user_choice("Do you want to load an Excel file with existing naming data? (y/n): ", ["y", "n"]) == "y"
    if config_data["load_excel"]:
        partial = load_data(load_excel=True, load_tei=False)

        if partial.get("excel") is not None:
            data["excel"] = partial["excel"]
            data["excel_path"] = partial.get("excel_path")
            config_data["excel_path"] = partial.get("excel_path")
        else:
            print("❌ No Excel file was loaded. Disabling Excel-related processing.")
            config_data["load_excel"] = False

    # 🔹 TEI laden (nach Eingabe)
    config_data["load_tei"] = input("Do you want to load the corresponding TEI file? (y/n): ").strip().lower() == "y"
    if config_data["load_tei"]:
        partial = load_data(load_excel=False, load_tei=True)

        if partial.get("xml") is not None:
            data["xml"] = partial["xml"]
            data["tei_path"] = partial.get("tei_path") or partial.get("xml_path")
            config_data["tei_path"] = data["tei_path"]
        else:
            print("❌ No TEI file was loaded. Disabling TEI-related processing.")
            config_data["load_tei"] = False

    print("Do you want to collect new data or analyze existing data?")
    print("[1] Collect data")
    print("[2] Analyze data")
    mode = ask_user_choice("> ", ["1", "2"])
    config_data["modus"] = "collect" if mode == "1" else "analyze"

    # ⛔ Skip data collection config if in analyze mode
    if config_data["modus"] == "analyze":
        save_config(config_path, config_data)
        return config_data, data

    config_data["check_namings"] = ask_user_choice("Should namings be checked and added? (y/n): ", ["y", "n"]) == "y"
    config_data["fill_collocations"] = ask_user_choice("Should empty collocations be filled? (y/n): ", ["y", "n"]) == "y"
    config_data["do_categorization"] = ask_user_choice("Should namings be lemmatized and categorized? (y/n): ", ["y", "n"]) == "y"

    save_config(config_path, config_data)

    return config_data, data

def save_config(path, config_data):
    try:
        safe_write_json(config_data, path)
        print(f"💾 Settings saved to: {path}")
    except Exception as e:
        print(f"❌ Failed to save config: {e}")

def run_data_collection(
    df,
    root,
    naming_dict,
    last_verse,
    paths,
    missing_naming_variants,
    collocation_data,
    check_namings=True,
    perform_collocations=False,
    perform_categorization=False,
    lemma_normalization=None,
    ignored_lemmas=None,
    lemma_categories=None,
    categorized_entries=None
):
    """
    Runs the data collection process depending on active modes:
    - If check_namings is True → TEI-based loop.
    - If only collocations and/or categorization are active → Excel-based loop.
    Returns updated (missing_naming_variants, collocation_data, categorized_entries).
    """

    verse = root.findall('.//tei:l', tei_ns)
    if not verse:
        print("⚠️ No verses found.")
        return missing_naming_variants, collocation_data, categorized_entries

    # --- TEI-based loop (only when check_namings is active)
    if check_namings:
        start_index = next(
            (i for i, line in enumerate(verse) if get_valid_verse_number(line.get("n")) > last_verse),
            0
        )

        print(f"🔁 Starting TEI iteration from verse {int(verse[start_index].get('n'))} (Index {start_index})")

        for line in verse[start_index:]:
            verse_number = get_valid_verse_number(line.get("n"))

            verse_text = ' '.join([seg.text for seg in line.findall(".//tei:seg", tei_ns) if seg.text])
            normalized_verse = normalize_text(verse_text)

            # Naming detection
            missing_naming_variants = check_and_extend_namings(
                verse_number,
                verse_text,
                normalized_verse,
                df,
                naming_dict,
                missing_naming_variants,
                root,
                paths,
                perform_categorization,
                lemma_normalization,
                ignored_lemmas,
                lemma_categories,
                categorized_entries
            )

            # Collocations
            if perform_collocations:
                rows = df[df["Vers"] == verse_number]

                for _, row in rows.iterrows():
                    check_and_add_collocations(
                        verse_number, collocation_data, root, paths, row=row
                    )

            # Categorization
            if perform_categorization:
                df_verse = df[(df["Vers"] >= verse_number) & (df["Vers"] < verse_number + 1)]
                entries = df_verse.to_dict(orient="records")

                for entry in entries:
                    source_text = normalize_text(get_first_valid_text(
                        entry.get("Erzähler"),
                        entry.get("Bezeichnung"),
                        entry.get("Eigennennung")
                    ))
                    if not source_text:
                        continue

                    skip = False
                    for e in categorized_entries:
                        if int(e.get("Vers", -1)) != verse_number:
                            continue

                        target_text = normalize_text(get_first_valid_text(
                            e.get("Erzähler"),
                            e.get("Bezeichnung"),
                            e.get("Eigennennung")
                        ))

                        if source_text == target_text and normalize_text(e.get("Benannte Figur", "")) == normalize_text(
                            entry.get("Benannte Figur", "")):
                            if any(
                                str(e.get(k, "")).strip()
                                for k in e.keys()
                                if k.startswith("Bezeichnung") or k.startswith("Epitheta")
                            ):
                                skip = True
                                break

                    if skip:
                        continue

                    annotated = lemmatize_and_categorize_entry(
                        entry, lemma_normalization, paths, ignored_lemmas, lemma_categories
                    )
                    if annotated:
                        categorized_entries.append(annotated)

            # Save progress after each verse
            save_progress(
                missing_naming_variants=missing_naming_variants,
                last_processed_verse=verse_number,
                paths=paths,
                check_namings=check_namings,
                perform_collocations=perform_collocations,
                perform_categorization=perform_categorization
            )

    # --- Excel-based loop (when only collocations and/or categorization are active)
    elif perform_collocations or perform_categorization:
        print("🔁 Starting EXCEL-based iteration over 'Vers' list.")

        # Extract and sort valid verse numbers from Excel
        vers_list = sorted(set(
            v for v in df["Vers"] if str(v).strip().isdigit()
        ))

        for v in vers_list:
            verse_number = get_valid_verse_number(v)

            # Collocations
            if perform_collocations:
                rows = df[df["Vers"] == verse_number]

                for _, row in rows.iterrows():
                    check_and_add_collocations(
                        verse_number, collocation_data, root, paths, row=row
                    )
            # Categorization
            if perform_categorization:
                df_verse = df[(df["Vers"] >= verse_number) & (df["Vers"] < verse_number + 1)]
                entries = df_verse.to_dict(orient="records")

                for entry in entries:
                    source_text = normalize_text(get_first_valid_text(
                        entry.get("Erzähler"),
                        entry.get("Bezeichnung"),
                        entry.get("Eigennennung")
                    ))
                    if not source_text:
                        continue

                    skip = False
                    for e in categorized_entries:
                        if int(e.get("Vers", -1)) != verse_number:
                            continue

                        target_text = normalize_text(get_first_valid_text(
                            e.get("Erzähler"),
                            e.get("Bezeichnung"),
                            e.get("Eigennennung")
                        ))

                        if source_text == target_text and normalize_text(e.get("Benannte Figur", "")) == normalize_text(
                            entry.get("Benannte Figur", "")):
                            if any(
                                str(e.get(k, "")).strip()
                                for k in e.keys()
                                if k.startswith("Bezeichnung") or k.startswith("Epitheta")
                            ):
                                skip = True
                                break

                    if skip:
                        continue

                    annotated = lemmatize_and_categorize_entry(
                        entry, lemma_normalization, paths, ignored_lemmas, lemma_categories
                    )
                    if annotated:
                        categorized_entries.append(annotated)

            # Save progress after each verse
            save_progress(
                missing_naming_variants=missing_naming_variants,
                last_processed_verse=verse_number,
                paths=paths,
                check_namings=check_namings,
                perform_collocations=perform_collocations,
                perform_categorization=perform_categorization
            )

    # Return updated data
    return missing_naming_variants, collocation_data, categorized_entries


def check_and_extend_namings(
    verse_number: int,
    verse_text: str,
    normalized_verse: str,
    df: pd.DataFrame,
    naming_dict: dict,
    missing_naming_variants: list,
    root: Element,
    paths: dict,
    perform_categorization: bool,
    lemma_normalization: dict,
    ignored_lemmas: set,
    lemma_categories: dict,
    categorized_entries: list
) -> list:
    """
    Checks whether a naming from the global dict appears in the current verse,
    but is not yet listed in Excel or in confirmed/rejected namings.
    If found: interactive confirmation and storage.
    """

    # 1. Extract namings from Excel for the current verse
    existing_namings = set()
    if "Vers" in df.columns:
        df_verse = df[df["Vers"] == verse_number]
        for column in ["Eigennennung", "Bezeichnung", "Erzähler"]:
            if column in df_verse.columns:
                values = df_verse[column].dropna().tolist()
                existing_namings.update(
                    normalize_text(str(value).strip()) for value in values if str(value).strip()
                )

    # 2. Extract and normalize namings from dict
    dict_namings = set()
    for book_list in naming_dict.get("Namings", {}).values():
        dict_namings.update(
            normalize_text(name.strip()) for name in book_list if name.strip()
        )

    # 3. Match check & user interaction
    for naming in dict_namings:
        if not naming:
            continue

        # skip if already handled in Excel (auch als Token-Menge)
        naming_tokens = set(naming.split())

        skip_existing = False
        for entry in existing_namings:
            entry_tokens = set(entry.split())
            if naming in entry or entry in naming:
                skip_existing = True
                break
            if naming_tokens <= entry_tokens or entry_tokens <= naming_tokens:
                skip_existing = True
                break

        if skip_existing:
            continue

        # skip if already handled in JSON
        skip = False
        for entry in missing_naming_variants:
            if entry.get("Vers") == verse_number:
                values = [
                    entry.get("Eigennennung", ""),
                    entry.get("Bezeichnung", ""),
                    entry.get("Erzähler", "")
                ]
                if normalize_text(naming) in map(normalize_text, values):
                    skip = True
                    break
        if skip:
            continue

        if not re.search(rf'\b{re.escape(naming)}\b', normalized_verse):
            continue

        print("\n" + "-" * 60)
        print(f"❗ New naming found that is not listed in the Excel file!")
        print(f"🔍 Detected naming: \"{naming}\"")

        # 📖 Show context
        prev_line = root.find(f'.//tei:l[@n="{verse_number - 1}"]', tei_ns)
        if prev_line is not None:
            prev_text = ' '.join([seg.text for seg in prev_line.findall('.//tei:seg', tei_ns) if seg.text])
            print(f"📖 Previous verse ({verse_number - 1}): {prev_text}")

        highlighted = verse_text.replace(naming, f"\033[1m\033[93m{naming}\033[0m")
        print(f"📖 Verse ({verse_number}): {highlighted}")

        next_line = root.find(f'.//tei:l[@n="{verse_number + 1}"]', tei_ns)
        if next_line is not None:
            next_text = ' '.join([seg.text for seg in next_line.findall('.//tei:seg', tei_ns) if seg.text])
            print(f"📖 Next verse ({verse_number + 1}): {next_text}")

        # 🧍 Confirm with user
        confirm = ask_user_choice("Is this a missing naming variant? (y/n): ", ["y", "n"])
        if confirm == "n":
            missing_naming_variants.append({
                "Vers": verse_number,
                "Eigennennung": naming,
                "Nennende Figur": "",
                "Bezeichnung": "",
                "Erzähler": "",
                "Status": "rejected"
            })
            save_progress(missing_naming_variants, verse_number, paths)
            print("✅ Rejection saved.")
            continue

        extend = ask_user_choice("💡 Would you like to shorten or lengthen the naming variant (y/n): ", ["y", "n"])
        if extend == "y":
            naming = input("✍ Enter the adapted naming variant: ").strip()

        print("Please choose the correct category:")
        print("[1] Eigennennung")
        print("[2] Bezeichnung")
        print("[3] Erzähler")
        print("[4] Skip")

        choice = input("👉 Your selection: ").strip()
        if choice == "4":
            continue

        named_entity = input("Enter the \"Benannte Figur\": ").strip()
        naming_entity = ""
        if choice == "2":
            naming_entity = input("Enter the \"Nennende Figur\": ").strip()

        entry = {
            "Benannte Figur": named_entity,
            "Vers": verse_number,
            "Eigennennung": naming if choice == "1" else "",
            "Nennende Figur": naming_entity,
            "Bezeichnung": naming if choice == "2" else "",
            "Erzähler": naming if choice == "3" else "",
            "Status": "confirmed"
        }

        # 📌 Optional: add collocation
        wants_collocation = ask_user_choice("📌 Do you want to add a collocation (context lines)? (y/n): ", ["y", "n"])
        if wants_collocation == "y":
            print("\n📖 Extended context (1–13):")
            context_lines = {}
            number = 1

            for i in range(6, 0, -1):
                line = root.find(f'.//tei:l[@n="{verse_number - i}"]', tei_ns)
                if line is not None:
                    text = ' '.join([seg.text for seg in line.findall('.//tei:seg', tei_ns) if seg.text])
                    context_lines[number] = text
                    print(f"[{number}] {text}")
                    number += 1

            context_lines[number] = verse_text
            print(f"[{number}] {verse_text}")
            number += 1

            for i in range(1, 7):
                line = root.find(f'.//tei:l[@n="{verse_number + i}"]', tei_ns)
                if line is not None:
                    text = ' '.join([seg.text for seg in line.findall('.//tei:seg', tei_ns) if seg.text])
                    context_lines[number] = text
                    print(f"[{number}] {text}")
                    number += 1

            selection = input("\n👉 Please enter the line number(s) (e.g., '5-7' or '6'): ").strip()
            selected = []

            try:
                if "-" in selection:
                    start, end = map(int, selection.split("-"))
                    selected = [context_lines[i] for i in range(start, end + 1) if i in context_lines]
                else:
                    idx = int(selection)
                    selected = [context_lines[idx]]
            except (ValueError, KeyError):
                print("⚠️ Invalid input – no collocation saved.")

            if selected:
                entry["Kollokation"] = ' / '.join(selected)

        missing_naming_variants.append(entry)
        save_progress(missing_naming_variants, verse_number, paths)
        print("✅ Entry saved.")

        # 🆕 Sofortige Kategorisierung, falls aktiviert und bestätigt
        if perform_categorization and entry["Status"] == "confirmed":
            annotated = lemmatize_and_categorize_entry(
                entry,
                lemma_normalization,
                paths,
                ignored_lemmas,
                lemma_categories
            )
            if annotated:
                categorized_entries.append(annotated)

    return missing_naming_variants

def load_missing_naming_variants(path: str) -> list:
    """
    Loads missing or confirmed namings from a JSON file.
    Returns an empty list if the file doesn't exist or is invalid.
    """
    return safe_read_json(path, default=[])


def get_verse_context(verse_number, root_tei):
    """Retrieves the surrounding 6 verses from the TEI file, numbered 1–13."""
    context = []
    verse_list = []

    for i in range(-6, 7):
        verse_id = str(verse_number + i)  # must be string!
        line = root_tei.find(f'.//tei:l[@n="{verse_id}"]', tei_ns)

        if line is not None:
            text = normalize_text(' '.join([
                seg.text for seg in line.findall('.//tei:seg', tei_ns) if seg.text
            ]))
            verse_list.append(text)

    for i, verse in enumerate(verse_list, start=1):
        context.append((i, verse))

    return context

def clean_cell_value(value):
    if pd.isna(value) or value is None:
        return ""
    return normalize_text(str(value).strip())


def sanitize_cell_value(value):
    """
    Cleans a cell value from invisible characters and special cases for robust empty-checking.
    """
    # Catch NA, None, and strange pandas 'nan' strings
    if pd.isna(value) or value is None or str(value).lower().strip() in {"", "nan", "na"}:
        return ""

    cleaned = str(value)
    # Remove invisible Unicode characters
    cleaned = re.sub(r'[\u200b\u200c\u200d\uFEFF\xa0]', '', cleaned)
    cleaned = cleaned.strip()
    return cleaned

def check_and_add_collocations(verse_number, collocation_data, root, paths, row):
    """Checks whether a collocation should be added – if so, prompts for user input."""

    # Check if already handled via Excel
    if sanitize_cell_value(row.get("Kollokationen")) != "":
        return None

    # Check if already handled via JSON
    if any(
        int(entry.get("Vers", -1)) == verse_number and str(entry.get("Kollokationen", "")).strip()
        for entry in collocation_data
    ):
        return None

    naming = clean_cell_value(row.get("Eigennennung")) \
             or clean_cell_value(row.get("Bezeichnung")) \
             or clean_cell_value(row.get("Erzähler"))

    named_entity = clean_cell_value(row.get("Benannte Figur"))

    context = get_verse_context(verse_number, root)

    collocations = ask_for_collocations(verse_number, named_entity, naming, context)

    collocation_data.append({
        "Vers": verse_number,
        "Kollokationen": collocations
    })

    # 📝 Immediately save progress
    with open(paths["collocations_json"], "w", encoding="utf-8") as f:
        json.dump(collocation_data, f, indent=4, ensure_ascii=False)

    return True

def ask_for_collocations(verse_number, named_entity, naming, context):
    """Displays the context around a verse and interactively asks for relevant collocations."""

    print(f"\n🟡 Empty collocation field detected in verse {verse_number}!")
    if named_entity or naming:
        print(f"👤 {named_entity}: {naming}\n")

    for number, text in context:
        if naming:
            # Highlight naming
            highlighted = text.replace(str(naming), f"\033[1;33m{naming}\033[0m")
        else:
            highlighted = text
        print(f"{number}. {highlighted}")

    user_input = input("\n👉 Please enter the number(s) of the relevant lines (e.g., '5' or '5-7'): ")

    selected = []

    try:
        if "-" in user_input:
            start, end = map(int, user_input.split("-"))
            selected = [text for number, text in context if start <= number <= end]
        else:
            number = int(user_input)
            selected = [text for num, text in context if num == number]
    except (ValueError, StopIteration):
        print("⚠️ Invalid input. Please enter a single number or a range.")
        return ""

    return " / ".join(selected)

def load_collocations_json(file_path):
    """Loads existing collocations from a JSON file – or returns an empty list if unavailable."""
    return safe_read_json(file_path, default=[])

def load_ignored_lemmas(path="ignored_lemmas.json"):
    data = safe_read_json(path, default=[])
    return set(data) if isinstance(data, list) else set(data.keys())

def load_lemma_categories(path="data/lemma_categories.json"):
    return safe_read_json(path, default={})

def load_json_annotations(path):
    return safe_read_json(path, default=[])

def save_json_annotations(path, annotations):
    safe_write_json(annotations, path, merge=True)

def resolve_lemma(token: str, lemma_dict: dict[str, list[str]]) -> str:
    """
    Finds the corresponding lemma for a given token using a dictionary
    of the form {lemma: [variants]}.
    Returns the token itself if no match is found (fallback).
    """
    for lemma, variants in lemma_dict.items():
        if token in variants:
            return lemma
    return token  # fallback if no variant matches

def lemmatize_and_categorize_entry(entry, lemma_normalization, paths, ignored_lemmas=None, lemma_categories=None):
    if lemma_normalization is None:
        lemma_normalization = load_lemma_normalization(paths["lemma_normalization_json"])

    if ignored_lemmas is None:
        ignored_lemmas = load_ignored_lemmas(paths["ignored_lemmas_json"])

    if lemma_categories is None:
        lemma_categories = load_lemma_categories(paths["lemma_categories_json"])

    text = get_first_valid_text(
        entry.get("Erzähler"),
        entry.get("Bezeichnung"),
        entry.get("Eigennennung")
    )

    if not text:
        print("⚠ No text to annotate – entry skipped.\n")
        return None

    print("\n" + "=" * 60)
    print(f"▶ Verse: {entry.get('Vers')}")
    print(f"▶ Named Entity: {entry.get('Benannte Figur')}")

    first_text = get_first_valid_text(
        entry.get("Eigennennung"),
        entry.get("Bezeichnung"),
        entry.get("Erzähler")
    )

    typ = "(unbestimmt)"

    if first_text == entry.get("Eigennennung"):
        typ = "Eigennennung"
    elif first_text == entry.get("Bezeichnung"):
        typ = "Bezeichnung"
    elif first_text == entry.get("Erzähler"):
        typ = "Erzähler"

    print(f"▶ Type: {typ}")

    print(f"\n▶ Original text: {text}")

    tokens = [t for t in tokenize(text.lower()) if t.isalpha()]

    # Filter only real word tokens
    missing = [
        t for t in tokens
        if t.isalpha() and not any(t in v or t == k for k, v in lemma_normalization.items())
    ]

    if missing:
        while True:
            print(f"\n▶ Please add lemma(ta) for {', '.join(missing)} (comma-separated):")
            user_input = input("> ").strip()
            new_lemmata = [l.strip() for l in user_input.split(",") if l.strip()]
            if len(new_lemmata) == len(missing):
                break
            print(
                f"⚠ Number of lemmata ({len(new_lemmata)}) doesn't match number of tokens ({len(missing)}). Please try again.")

        for token, lemma in zip(missing, new_lemmata):
            lemma_normalization.setdefault(lemma, [])
            if token not in lemma_normalization[lemma]:
                lemma_normalization[lemma].append(token)

        # 🔤 Sort alphabetically
        for lemma in lemma_normalization:
            lemma_normalization[lemma] = sorted(set(lemma_normalization[lemma]))

        save_lemma_normalization(lemma_normalization, path=paths["lemma_normalization_json"])

    lemmata = [resolve_lemma(t, lemma_normalization) for t in tokens]
    print(f"\n▶ Lemma: {', '.join(lemmata)}\n")

    while True:
        designations, epithets = run_categorization(
            lemmata, lemma_categories, ignored_lemmas, paths
        )

        if not designations and not epithets:
            print("⚠ No entry – please review and confirm again.")
            confirm = ask_user_choice("Really skip this entry? [y = yes / n = no]: ", ["y", "n"])
            if confirm == "y":
                print("⏭ Entry skipped.\n")
                return None
            else:
                return run_categorization(lemmata, lemma_categories, ignored_lemmas, paths)
        else:
            break

    if lemma_normalization:
        save_lemma_normalization(lemma_normalization, path=paths["lemma_normalization_json"])

    if ignored_lemmas:
        save_ignored_lemmas(ignored_lemmas, path=paths["ignored_lemmas_json"])

    if lemma_categories:
        save_lemma_categories(lemma_categories, path=paths["lemma_categories_json"])

    annotated_entry = {
        **entry,
        "Bezeichnung 1": designations[0] if len(designations) > 0 else "",
        "Bezeichnung 2": designations[1] if len(designations) > 1 else "",
        "Bezeichnung 3": designations[2] if len(designations) > 2 else "",
        "Bezeichnung 4": designations[3] if len(designations) > 3 else "",
        "Epitheta 1": epithets[0] if len(epithets) > 0 else "",
        "Epitheta 2": epithets[1] if len(epithets) > 1 else "",
        "Epitheta 3": epithets[2] if len(epithets) > 2 else "",
        "Epitheta 4": epithets[3] if len(epithets) > 3 else "",
        "Epitheta 5": epithets[4] if len(epithets) > 4 else ""
    }

    # 💾 Kategorisierung direkt speichern
    existing = load_json_annotations(paths["categorization_json"])
    existing.append(annotated_entry)
    safe_write_json(existing, paths["categorization_json"], merge=True)

    print("✅ Entry saved.\n")
    return annotated_entry

def run_categorization(lemmata, lemma_categories, ignored_lemmas, paths):
    while True:
        designations = []
        epithets = []
        history = []
        i = 0

        while i < len(lemmata):
            lemma = lemmata[i]

            if lemma in ignored_lemmas:
                i += 1
                continue

            default = f"[{lemma_categories.get(lemma, '')}]" if lemma in lemma_categories else ""
            print(f"{lemma:<12} → {default} ", end="")
            user_input = input().strip()

            if user_input == "<":
                if i == 0 or not history:
                    print("↩️  Already at beginning – can't step back.")
                    continue
                i -= 1
                last_action = history.pop()
                if last_action["type"] == "a":
                    designations.pop()
                elif last_action["type"] == "e":
                    epithets.pop()
                elif last_action["type"] == "ignore":
                    ignored_lemmas.discard(last_action["lemma"])
                    save_ignored_lemmas(ignored_lemmas, path=paths["ignored_lemmas_json"])
                elif last_action["type"] == "override":
                    del lemma_categories[last_action["lemma"]]
                    save_lemma_categories(lemma_categories, path=paths["lemma_categories_json"])
                continue

            if user_input == "" and default:
                if default == "[a]":
                    designations.append(lemma)
                    history.append({"type": "a", "lemma": lemma})
                elif default == "[e]":
                    epithets.append(lemma)
                    history.append({"type": "e", "lemma": lemma})
                i += 1
                continue

            if user_input == "":
                confirm_ignore = ask_user_choice(f"⚠️ Really ignore lemma “{lemma}”? [y/n]: ", ["y", "n"])
                if confirm_ignore == "y":
                    ignored_lemmas.add(lemma)
                    save_ignored_lemmas(ignored_lemmas, path=paths["ignored_lemmas_json"])
                    print(f"ℹ️ Lemma “{lemma}” added to ignore list.")
                    history.append({"type": "ignore", "lemma": lemma})
                    i += 1
                    continue
                else:
                    print("↩️  Skipped ignoring – please choose a category or go back.\n")
                    continue

            if user_input in ("a", "e"):
                if user_input == "a":
                    designations.append(lemma)
                else:
                    epithets.append(lemma)
                lemma_categories[lemma] = user_input
                save_lemma_categories(lemma_categories, path=paths["lemma_categories_json"])
                history.append({"type": user_input, "lemma": lemma})
                i += 1
                continue

            correction = user_input
            cat = ""
            while cat not in ("a", "e"):
                cat = input(f'Define category for “{correction}” [a/e]: ').strip().lower()

            if cat == "a":
                designations.append(correction)
            else:
                epithets.append(correction)

            lemma_categories[correction] = cat
            save_lemma_categories(lemma_categories)
            history.append({"type": "override", "lemma": correction})
            i += 1

        return designations, epithets

def tokenize(text):
    return re.findall(r'\w+|[^\w\s]', text, re.UNICODE)

def load_lemma_normalization(path="lemma_normalization.json"):
    return safe_read_json(path, default={})

def save_lemma_normalization(data, path="lemma_normalization.json"):
    sorted_data = {
        lemma: sorted(set(variants))
        for lemma, variants in sorted(data.items(), key=lambda x: x[0].lower())
    }
    safe_write_json(sorted_data, path, merge=False)

def save_ignored_lemmas(data, path="ignored_lemmas.json"):
    safe_write_json(data, path, sort_keys=True, merge=True)

def save_lemma_categories(data, path="data/lemma_categories.json"):
    existing = safe_read_json(path, default={})
    existing.update(data)
    sorted_data = dict(sorted(existing.items()))
    safe_write_json(sorted_data, path, merge=False)

def run_analysis_menu(config_data, paths, data, book_name):
    while True:
        print("📊 Which type of analysis do you want to perform?")
        print("[1] Wordlist")
        print("[2] Keywords")
        print("[3] Collocations")
        print("[4] Exit analysis")

        choice = ask_user_choice("> ", ["1", "2", "3", "4"])

        if choice == "1":
            run_wordlist_menu(paths, book_name)
        elif choice == "2":
            run_keyword_menu(config_data, paths, data, book_name)
        elif choice == "3":
            run_collocation_menu(config_data, paths, data, book_name)
        elif choice == "4":
            print("📦 Analysis completed.")
            break

def run_wordlist_menu(paths, book_name):
    """
    Interactive menu for generating wordlists:
    - by column
    - by figure (Bezeichnungen, Epitheta, combined)
    Results are saved as CSV files.
    """

    json_path = paths["categorization_json"]
    output_dir = os.path.join("data", book_name, "analysis")
    os.makedirs(output_dir, exist_ok=True)

    while True:
        print("\n📁 What kind of wordlist do you want to generate?")
        print("[1] All values from a column (e.g., 'Benannte Figur')")
        print("[2] All designations (Bezeichnungen) for a specific figure")
        print("[3] All epithets (Epitheta) for a specific figure")
        print("[4] Combined designations and epithets")
        print("[5] Back to main analysis menu")

        choice = ask_user_choice("> ", ["1", "2", "3", "4", "5"])

        if choice == "1":
            print("\n📑 Available column groups:")
            print("- Benannte Figur")
            print("- Bezeichnung")
            print("- Epitheta")

            column_input = input("Please enter one of the above column names:\n> ").strip()
            valid_columns = ["Benannte Figur", "Bezeichnung", "Epitheta"]

            while column_input not in valid_columns:
                print("⚠️ Invalid input. Please enter one of: Benannte Figur, Bezeichnung, Epitheta")
                column_input = input("> ").strip()

            filename = f"wordlist_{column_input}_{book_name}.csv".replace(" ", "_")
            output_path = os.path.join(output_dir, filename)
            generate_wordlist_by_column(column_input, json_path, output_path)

        elif choice == "2":
            figure = ask_valid_figure_name(paths["categorization_json"])
            if figure is None:
                return
            filename = f"wordlist_Bezeichnung_{figure}.csv".replace(" ", "_")
            output_path = os.path.join(output_dir, filename)
            generate_designations_for_figure(figure, json_path, output_path)

        elif choice == "3":
            figure = ask_valid_figure_name(paths["categorization_json"])
            if figure is None:
                return
            filename = f"wordlist_Epitheta_{figure}.csv".replace(" ", "_")
            output_path = os.path.join(output_dir, filename)
            generate_epithets_for_figure(figure, json_path, output_path)

        elif choice == "4":
            figure = ask_valid_figure_name(paths["categorization_json"])
            if figure is None:
                return
            filename = f"wordlist_Combined_{figure}.csv".replace(" ", "_")
            output_path = os.path.join(output_dir, filename)
            generate_combined_designations_epithets(figure, json_path, output_path)

        elif choice == "5":
            print("↩️ Returning to analysis menu.")
            return

def generate_wordlist_by_column(column_name: str, json_path: str, output_path: str):
    """
    Creates a frequency list for the selected column (or column group) from a categorization JSON file.
    Saves the result as a CSV file.

    :param column_name: Name of the column or logical group (e.g., "Bezeichnung", "Epitheta", "Benannte Figur")
    :param json_path: Path to categorization_<book>.json
    :param output_path: Target path for the CSV file
    """
    entries = safe_read_json(json_path, default=[])

    if column_name.lower() == "bezeichnung":
        columns = [f"Bezeichnung {i}" for i in range(1, 5)]
    elif column_name.lower() == "epitheta":
        columns = [f"Epitheta {i}" for i in range(1, 6)]
    else:
        columns = [column_name]

    all_values = []
    for entry in entries:
        for col in columns:
            value = entry.get(col)
            if isinstance(value, str) and value.strip():
                all_values.append(value.strip())

    counts = Counter(all_values)
    most_common = counts.most_common()

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Wert", "Anzahl"])
        for value, count in most_common:
            writer.writerow([value, count])

    print(f"✅ Wordlist written to: {output_path}")

def resolve_figure_name(name: str, entries: list[dict]) -> str | None:
    """
    Checks if the given figure name exists in the data. If not, suggests the closest match.
    Returns the resolved name or None if the user rejects the suggestion or no match is found.
    """
    all_names = {
        str(name).strip()
        for name in [e.get("Benannte Figur") for e in entries]
        if isinstance(name, str) and name.strip()
    }
    if name in all_names:
        return name

    suggestions = difflib.get_close_matches(name, all_names, n=1, cutoff=0.6)
    if suggestions:
        print(f'⚠️ Figure "{name}" not found.')
        print(f'❓ Did you mean "{suggestions[0]}"? [y/n]')
        answer = ask_user_choice("> ", ["y", "n"])
        if answer == "y":
            return suggestions[0]
        else:
            print("⚠️ No valid figure selected.")
            print("Please enter a valid name exactly as it appears in your categorization data.")
            return None
    else:
        print(f'⚠️ Figure "{name}" not found and no similar name could be suggested.')
        return None

def ask_valid_figure_name(json_path: str) -> str | None:
    entries = safe_read_json(json_path, default=[])

    while True:
        raw = input("✍ Please enter the figure name:\n> ").strip()
        if not raw:
            print("⚠️ Input cannot be empty.")
            continue

        resolved = resolve_figure_name(raw, entries)
        if resolved is not None:
            return resolved

        print("⚠️ No matching figure found. Please try again.")

    return None  # ← für statische Typprüfung, wird nie erreicht



def generate_designations_for_figure(figure_name: str, json_path: str, output_path: str):
    entries = safe_read_json(json_path, default=[])
    # no need to resolve again – already handled
    resolved_name = figure_name

    filtered = [e for e in entries if e.get("Benannte Figur") == resolved_name]
    values = []
    for entry in filtered:
        for i in range(1, 5):
            val = entry.get(f"Bezeichnung {i}")
            if isinstance(val, str) and val.strip():
                values.append(val.strip())

    counts = Counter(values)
    most_common = counts.most_common()

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Wert", "Anzahl"])
        for val, count in most_common:
            writer.writerow([val, count])

    print(f"✅ Wordlist for '{resolved_name}' written to: {output_path}")

def generate_epithets_for_figure(figure_name: str, json_path: str, output_path: str):
    entries = safe_read_json(json_path, default=[])
    # no need to resolve again – already handled
    resolved_name = figure_name

    filtered = [e for e in entries if e.get("Benannte Figur") == resolved_name]
    values = []
    for entry in filtered:
        for i in range(1, 6):
            val = entry.get(f"Epitheta {i}")
            if isinstance(val, str) and val.strip():
                values.append(val.strip())

    counts = Counter(values)
    most_common = counts.most_common()

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Wert", "Anzahl"])
        for val, count in most_common:
            writer.writerow([val, count])

    print(f"✅ Wordlist for epithets of '{resolved_name}' written to: {output_path}")

def generate_combined_designations_epithets(figure_name: str, json_path: str, output_path: str):
    """
    Generates a combined wordlist of all designations and epithets for a given figure.
    Bezeichnung 1–4 and Epitheta 1–5 are combined and counted together.
    Output is saved as CSV with 'Wert' and 'Anzahl'.
    """
    entries = safe_read_json(json_path, default=[])
    # no need to resolve again – already handled
    resolved_name = figure_name

    filtered = [e for e in entries if e.get("Benannte Figur") == resolved_name]
    values = []
    for entry in filtered:
        for i in range(1, 5):
            val = entry.get(f"Bezeichnung {i}")
            if isinstance(val, str) and val.strip():
                values.append(val.strip())
        for i in range(1, 6):
            val = entry.get(f"Epitheta {i}")
            if isinstance(val, str) and val.strip():
                values.append(val.strip())

    counts = Counter(values)
    most_common = counts.most_common()

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Wert", "Anzahl"])
        for val, count in most_common:
            writer.writerow([val, count])

    print(f"✅ Combined wordlist for '{resolved_name}' written to: {output_path}")

def run_keyword_menu(config_data, paths, data, book_name):
    """
    Interactive menu for configuring and running a keyword analysis.
    Collects target, reference, unit and threshold, then calls generate_keywords().
    """

    target_json = paths["categorization_json"]
    output_dir = os.path.join("data", book_name, "analysis")
    os.makedirs(output_dir, exist_ok=True)

    print("\n📌 Do you want to analyze the whole work or a specific figure?")
    print("[1] Whole work")
    print("[2] Specific figure")

    target_choice = ask_user_choice("> ", ["1", "2"])

    reference_books = None

    if target_choice == "2":
        target = ask_valid_figure_name(target_json)
        if target is None:
            return None
        target_type = "figure"

    else:
        target = book_name
        target_type = "whole_work"
        print("📘 Please enter the names of the works to include in the reference corpus (comma-separated):")
        references = input("> ").strip()
        reference_books = [r.strip() for r in references.split(",") if r.strip()]

    print("\n🎯 What should be the unit of comparison?")
    print("[1] Designations (Bezeichnungen)")
    print("[2] Epithets (Epitheta)")
    print("[3] Combined")

    unit_choice = ask_user_choice("> ", ["1", "2", "3"])
    unit = {
        "1": "bezeichnung",
        "2": "epitheta",
        "3": "combined"
    }[unit_choice]

    print("\n🧪 Enter significance threshold (Log-Likelihood G²), default = 3.84:")
    threshold_input = input("> ").strip()
    try:
        threshold = float(threshold_input) if threshold_input else 3.84
    except ValueError:
        print("⚠️ Invalid input – using default threshold 3.84")
        threshold = 3.84

    # Prepare output filename
    target_label = target.replace(" ", "_")
    output_file = f"keywords_{unit}_{target_label}_{book_name}.csv"
    output_path = os.path.join(output_dir, output_file)

    # Call the actual keyword function
    if target_type == "figure":
        generate_keywords(
            target_figure=target,
            reference_books=None,
            unit=unit,
            threshold=threshold,
            target_json=target_json,
            output_path=output_path
        )
    else:
        generate_keywords(
            target_figure=None,
            reference_books=reference_books,
            unit=unit,
            threshold=threshold,
            target_json=target_json,
            output_path=output_path
        )

    print(f"✅ Keyword analysis written to: {output_path}")

    print("\n🔁 Do you want to run another keyword analysis? [y/n]")
    again = ask_user_choice("> ", ["y", "n"])
    if again == "y":
        return run_keyword_menu(config_data, paths, data, book_name)
    else:
        print("↩️ Returning to analysis menu.")
        return None

def generate_keywords(
    target_figure: str | None,
    reference_books: list[str] | None,
    unit: str,
    threshold: float,
    target_json: str,
    output_path: str
):
    """
    Calculates key terms (designations, epithets or both) for a figure or full work
    compared to a reference corpus. Saves results as CSV (only values above threshold).
    """
    target_entries = safe_read_json(target_json, default=[])

    # Filter target corpus
    if target_figure:
        target_entries = [e for e in target_entries if e.get("Benannte Figur") == target_figure]

    target_tokens = extract_tokens(target_entries, unit)

    # Load reference corpus
    reference_entries = []

    if reference_books:
        for book in reference_books:
            path = os.path.join("data", f"categorization_{book}.json")
            reference_entries += safe_read_json(path, default=[])
    else:
        # fallback: all entries except target_figure
        reference_entries = [
            e for e in safe_read_json(target_json, default=[])
            if not target_figure or e.get("Benannte Figur") != target_figure
        ]

    reference_tokens = extract_tokens(reference_entries, unit)

    # Count occurrences
    target_counts = Counter(target_tokens)
    reference_counts = Counter(reference_tokens)

    results = []
    total_target = sum(target_counts.values())
    total_ref = sum(reference_counts.values())

    for token, count_t in target_counts.items():
        count_r = reference_counts.get(token, 0)

        if count_t + count_r == 0:
            continue

        # Log-likelihood Berechnung (G²)
        p = (count_t + count_r) / (total_target + total_ref)
        expected_t = p * total_target
        expected_r = p * total_ref

        log_t = count_t * math.log2(count_t / expected_t) if count_t > 0 and expected_t > 0 else 0
        log_r = count_r * math.log2(count_r / expected_r) if count_r > 0 and expected_r > 0 else 0

        keyness = 2 * (log_t + log_r)

        if keyness >= threshold:
            if count_t > count_r:
                typ = "positive"
            elif count_r > count_t:
                typ = "negative"
            else:
                typ = "neutral"

            results.append((token, count_t, count_r, round(keyness, 2), typ))

    # Sort descending
    results.sort(key=lambda x: (-x[3], x[0]))

    # Write CSV
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Wort", "Zielanzahl", "Referenzanzahl", "Keyness", "Typ"])
        for row in results:
            writer.writerow(row)


def extract_tokens(entries: list[dict], unit: str) -> list[str]:
    """
    Extracts all relevant tokens (Bezeichnungen, Epitheta, or both) from a list of entries.

    :param entries: list of JSON objects
    :param unit: "bezeichnung", "epitheta", or "combined"
    :return: flat list of strings
    """
    tokens = []

    for entry in entries:
        if unit in ("bezeichnung", "combined"):
            for i in range(1, 5):
                val = entry.get(f"Bezeichnung {i}")
                if isinstance(val, str) and val.strip():
                    tokens.append(val.strip())

        if unit in ("epitheta", "combined"):
            for i in range(1, 6):
                val = entry.get(f"Epitheta {i}")
                if isinstance(val, str) and val.strip():
                    tokens.append(val.strip())

    return tokens

def run_collocation_menu(config_data, paths, data, book_name):
    """
    Interactive menu to search for collocation contexts of a specific type
    (designation or epithet), optionally restricted to a figure.
    Results can be shown in the console or saved as CSV.
    """

    _ = config_data
    categorization_path = paths["categorization_json"]

    print("\n📌 Do you want to analyze the whole work or only a specific figure?")
    print("[1] Whole work")
    print("[2] Specific figure")

    target_mode = ask_user_choice("> ", ["1", "2"])
    only_figure = None

    if target_mode == "2":
        only_figure = ask_valid_figure_name(categorization_path)
        if only_figure is None:
            return

    type_value = input("🔍 Please enter the type to search for (e.g., \"küene\"):\n> ").strip()

    while True:
        print("\n📤 Where should the results be displayed?")
        print("[1] Console")
        print("[2] Save as CSV file")

        output_choice = ask_user_choice("> ", ["1", "2"])
        output_target = "console" if output_choice == "1" else "csv"

        if output_target == "csv":
            type_label = type_value.replace(" ", "_")
            fig_label = only_figure.replace(" ", "_") if only_figure else "whole_work"
            output_dir = os.path.join("data", book_name, "analysis")
            os.makedirs(output_dir, exist_ok=True)
            filename = f"collocations_{fig_label}_{type_label}_{book_name}.csv"
            output_path = os.path.join(output_dir, filename)
        else:
            output_path = None

        try:
            generate_collocations(
                data=data,
                type_value=type_value,
                book_name=book_name,
                config_data=config_data,
                only_figure=only_figure,
                output_target=output_target,
                output_path=output_path
            )
            break  # ✅ innerhalb von while

        except PermissionError:
            print("\n⚠️ The Excel file appears to be open.")
            print("❗ Please close it and try again.")
            print("↩️ Returning to output choice...\n")


def generate_collocations(
    data: dict,
    type_value: str,
    book_name: str,
    config_data: dict,
    only_figure: str | None,
    output_target: str,
    output_path: str | None
):
    """
    Extracts all collocations containing a given type string from the final Excel file,
    optionally filtered by figure. Matches type occurrences from the JSON categorization file.
    Outputs formatted KWIC lines either to console or as CSV.
    """
    json_path = os.path.join("data", book_name, f"categorization_{book_name}.json")
    entries = safe_read_json(json_path, default=[])
    lemma_map = safe_read_json("data/lemma_normalization.json", default={})

    # Filter entries by figure if given
    if only_figure:
        entries = [e for e in entries if e.get("Benannte Figur") == only_figure]

    # Load Excel with fallback and sheet/column check
    df = load_collocation_sheet(config_data, book_name)
    if df is None:
        print("⚠️ Could not load the Excel sheet with 'Kollokationen'.")
        print("🔄 Falling back to TEI to reconstruct collocations.")
        df = build_fallback_collocation_df_from_tei(data["xml"])

    results = []

    for entry in entries:
        all_type_fields = [
            entry.get(f"Bezeichnung {i}") for i in range(1, 5)
        ] + [
            entry.get(f"Epitheta {i}") for i in range(1, 6)
        ]

        if not any(t == type_value for t in all_type_fields if isinstance(t, str)):
            continue

        vers = entry.get("Vers")
        figur = entry.get("Benannte Figur")
        original_text = get_first_valid_text(
            entry.get("Erzähler"),
            entry.get("Bezeichnung"),
            entry.get("Eigennennung")
        )

        match = df[
            (df["Vers"] == vers) &
            (df["Benannte Figur"] == figur) &
            (df.apply(lambda r: get_first_valid_text(
                r.get("Erzähler"),
                r.get("Bezeichnung"),
                r.get("Eigennennung")
            ) == original_text, axis=1))
        ]

        if match.empty:
            continue

        kollokation = match.iloc[0].get("Kollokationen")
        if not isinstance(kollokation, str) or not kollokation.strip():
            continue

                # Hole alle zugehörigen Varianten aus dem Lemma-Mapping
        raw_variants = lemma_map.get(type_value, [])

        # Sicherheit: Nur gültige Strings verwenden
        variants: List[str] = [v.strip() for v in raw_variants if isinstance(v, str) and v.strip()]

        # Fallback: Original-Keyword selbst auch aufnehmen (kleingeschrieben)
        if isinstance(type_value, str) and type_value.strip():
            variants.append(type_value.strip().lower())

        left, hit, right = format_kwic(kollokation, variants)
        results.append((vers, figur, left, hit, right))

    # Output formatting
    if output_target == "console":
        for _, _, left, hit, right in results:
            print(f"{left.strip():>40}  \033[1m\033[93m{hit}\033[0m  {right.strip():<40}")
    elif output_target == "csv" and output_path:
        with open(output_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["Vers", "Benannte Figur", "Left", "Hit", "Right"])
            for row in results:
                writer.writerow(row)
        print(f"✅ Collocation results saved to: {output_path}")

def load_collocation_sheet(config_data: dict, book_name: str) -> pd.DataFrame | None:
    """
    Loads the Excel sheet 'Gesamt' from either the finalized file in data/<book>/<book>_final.xlsx
    or from the fallback path stored in config_data["excel_path"].
    Returns the DataFrame if successful, else None.
    Raises PermissionError if the file is open and cannot be read.
    """
    primary_path = os.path.join("data", book_name, f"{book_name}_final.xlsx")
    fallback_path = config_data.get("excel_path")

    if os.path.exists(primary_path):
        try:
            df = pd.read_excel(primary_path, sheet_name="Gesamt", engine="openpyxl")
            if "Kollokationen" not in df.columns:
                print(f"⚠️ Sheet 'Gesamt' in file '{primary_path}' has no 'Kollokationen' column.")
                return None
            return df
        except PermissionError as e:
            raise e  # lassen wir weiter oben abfangen

    # Nur wenn Datei nicht existiert: Fallback
    if fallback_path and os.path.exists(fallback_path):
        try:
            df = pd.read_excel(fallback_path, sheet_name="Gesamt", engine="openpyxl")
            if "Kollokationen" not in df.columns:
                print(f"⚠️ Sheet 'Gesamt' in fallback file '{fallback_path}' has no 'Kollokationen' column.")
                return None
            return df
        except Exception as e:
            print(f"⚠️ Could not load fallback file: {e}")
            return None

    print("❌ No valid Excel file found for collocations.")
    return None

def format_kwic(context: str, variants: list[str]) -> tuple[str, str, str]:
    """
    Splits the context string into left, hit (matched variant), and right parts.
    Only the first match from the list of variants is considered.
    """
    context_lower = context.lower()

    for variant in variants:
        index = context_lower.find(variant.lower())
        if index != -1:
            left = context[:index].strip()
            hit = context[index:index + len(variant)]
            right = context[index + len(variant):].strip()
            return left, hit, right

    # No variant matched
    return context.strip(), "", ""

def build_fallback_collocation_df_from_tei(root_tei: Element) -> pd.DataFrame:
    """
    Constructs a DataFrame that simulates the Excel 'Kollokationen'-sheet,
    using ±3 verse context from the TEI structure.
    """
    context_data = []
    verses = root_tei.findall('.//tei:l', tei_ns)

    for idx, line in enumerate(verses):
        n_attr = line.get("n")
        if not n_attr or not n_attr.isdigit():
            continue
        verse_num = int(n_attr)

        # Kontext: Verse v-3 bis v+3
        segment_texts = []
        for offset in range(-3, 4):
            target_idx = idx + offset
            if 0 <= target_idx < len(verses):
                target_line = verses[target_idx]
                segs = [seg.text for seg in target_line.findall('.//tei:seg', tei_ns) if seg.text]
                segment_texts.append(" ".join(segs))

        full_context = " / ".join(segment_texts)
        context_data.append({"Vers": verse_num, "Kollokationen": full_context})

    return pd.DataFrame(context_data)

def export_all_data_to_new_excel(book_name, paths, options):
    """
    Integrates confirmed namings, adds collocations, and creates a lemmatized worksheet.

    :param book_name: The name of the text corpus (used for directory and filename construction)
    :param paths: Dictionary containing file paths (e.g., original_excel, missing_naming_variants_json, etc.)
    :param options: Dictionary with Boolean flags for: benennungen, kollokationen, kategorisierung
    """

    print("🟢 Starting export of all naming data...")

    # Support alternate keys for export paths
    paths = {
        **paths,
        "json_benennungen": paths.get("json_benennungen") or paths.get("missing_naming_variants_json"),
        "json_kollokationen": paths.get("json_kollokationen") or paths.get("collocations_json"),
        "json_kategorisierung": paths.get("json_kategorisierung") or paths.get("categorization_json"),
    }

    # 🔧 Ensure output directory exists
    project_dir = os.path.join("data", book_name)
    os.makedirs(project_dir, exist_ok=True)
    target_path = os.path.join(project_dir, f"{book_name}_final.xlsx")

    # Copy Excel file
    shutil.copy(paths["original_excel"], target_path)

    wb = openpyxl.load_workbook(target_path)
    sheet = wb["Gesamt"]

    if options.get("benennungen", False):
        print("📤 Exporting confirmed naming variants...")
        insert_naming_variants(sheet, paths["missing_naming_variants_json"])

    if options.get("kollokationen", False):
        print("📤 Exporting collocations...")
        update_collocations(sheet, paths["collocations_json"])

    if options.get("kategorisierung", False):
        print("📤 Exporting categorized lemmata (this may take a second)...")
        create_categorized_lemmas_sheet(wb, sheet, paths["categorization_json"])

    wb.save(target_path)
    print(f"✅ Export completed: {target_path}")

def get_format_template(sheet, column_index):
    """
    Returns formatting template (font, alignment, border, number format) of the first filled cell in the column.
    """
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=column_index)
        if cell.value:
            if cell.has_style:
                return copy(cell.font), copy(cell.alignment), copy(cell.border), cell.number_format
    return None, None, None, None

def insert_naming_variants(sheet, json_path):
    """
    Inserts confirmed namings into the 'Gesamt' worksheet.
    Column formatting is inherited from 'get_format_template()'.
    New rows are visually highlighted.
    """

    data = safe_read_json(json_path, default=[])

    confirmed_entries = [entry for entry in data if entry.get("Status") == "confirmed"]
    if not confirmed_entries:
        print("ℹ️ No confirmed naming variants to insert.")
        return

    last_line = sheet.max_row + 1
    fill_color = PatternFill(start_color="4BACC6", end_color="4BACC6", fill_type="solid")

    for entry in confirmed_entries:
        new_line = [
            entry.get("Benannte Figur", ""),
            entry.get("Vers", ""),
            entry.get("Eigennennung", ""),
            entry.get("Nennende Figur", ""),
            entry.get("Bezeichnung", ""),
            entry.get("Erzähler", ""),
            entry.get("Kollokation", "")
        ]

        for col_num, value in enumerate(new_line, start=1):
            cell = sheet.cell(row=last_line, column=col_num, value=value)

            font_tpl, alignment_tpl, border_tpl, number_format_tpl = get_format_template(sheet, col_num)
            if font_tpl:
                cell.font = font_tpl
                cell.alignment = alignment_tpl
                cell.border = border_tpl
                cell.number_format = number_format_tpl

            cell.fill = fill_color

        last_line += 1

    print("✅ Naming variants successfully added.")

def update_collocations(sheet, json_path):
    """
    Updates the 'Kollokationen' column using the JSON data.
    Formatting is copied from the first filled cell via get_format_template().
    """
    data = safe_read_json(json_path, default=[])

    header = [cell.value for cell in sheet[1]]
    try:
        verse_col = header.index("Vers") + 1
        collocation_col = header.index("Kollokationen") + 1
    except ValueError:
        print("❌ Columns 'Vers' or 'Kollokationen' not found!")
        return

    verse_to_rows = {}
    for row in range(2, sheet.max_row + 1):
        verse_value = sheet.cell(row=row, column=verse_col).value
        if verse_value is not None:
            verse_to_rows.setdefault(int(verse_value), []).append(row)

    font_tpl, alignment_tpl, border_tpl, number_format_tpl = get_format_template(sheet, collocation_col)

    updated_count = 0
    for entry in data:
        verse = entry["Vers"]
        new_value = entry["Kollokationen"]
        matching_rows = verse_to_rows.get(verse, [])
        for row in matching_rows:
            cell = sheet.cell(row=row, column=collocation_col, value=new_value)
            if font_tpl:
                cell.font = font_tpl
                cell.alignment = alignment_tpl
                cell.border = border_tpl
                cell.number_format = number_format_tpl
            updated_count += 1

    print(f"✅ {updated_count} collocations successfully updated.")

def create_categorized_lemmas_sheet(wb, _, json_path):
    """
    Creates a new worksheet 'lemmatisiert' with structured designations and epithets.
    Format and structure are copied from the 'Gesamt' sheet in the template.
    """

    # Load JSON data
    annotations = safe_read_json(json_path, default=[])

    # Replace existing sheet if necessary
    if "lemmatisiert" in wb.sheetnames:
        del wb["lemmatisiert"]
    ws_new = wb.create_sheet("lemmatisiert")

    # Define base styles
    regular_font = Font(name="Times New Roman", size=8, bold=False)
    bold_font = Font(name="Times New Roman", size=8, bold=True)

    default_alignment = Alignment(horizontal="left", vertical="bottom")  # bottom-aligned
    default_border = Border()

    # Define column headers
    headers = [
        "Benannte Figur", "Vers", "Eigennennung", "Nennende Figur", "Bezeichnung", "Erzähler",
        "Bezeichnung 1", "Bezeichnung 2", "Bezeichnung 3", "Bezeichnung 4",
        "Epitheta 1", "Epitheta 2", "Epitheta 3", "Epitheta 4", "Epitheta 5"
    ]

    df = pd.DataFrame(annotations)
    for col in headers:
        if col not in df.columns:
            df[col] = ""
    df = df[headers]

    # Write header row with bold formatting
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        cell = ws_new.cell(row=1, column=col_idx, value=header)
        ws_new.column_dimensions[col_letter].width = 20
        cell.font = bold_font
        cell.alignment = default_alignment
        cell.border = default_border
        cell.number_format = "General"

    # Write data rows with regular formatting
    for row_idx, row in df.iterrows():
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_new.cell(row=row_idx + 2, column=col_idx, value=row[header])
            cell.font = regular_font
            cell.alignment = default_alignment
            cell.border = default_border
            cell.number_format = "General"

    # Freeze only the first row
    ws_new.freeze_panes = "A2"

    # Enable auto-filter on the header
    ws_new.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Move the sheet to second position
    wb._sheets.insert(1, wb._sheets.pop(wb._sheets.index(ws_new)))

    print("✅ Worksheet 'lemmatisiert' successfully created.")

def main():
    # 🔹 1. Initialization: book selection, paths, last verse
    book_name, namings_last_verse, collocations_last_verse, categorization_last_verse, paths = initialize_project()

    # 🔹 2. Load global naming dictionary (from all books)
    naming_dict = load_or_extend_naming_dict()

    # 🔹 3. Konfiguration abfragen (interaktiv oder reuse)
    config_data, data = ask_config_interactively(paths["config_json"])

    paths["original_excel"] = data.get("excel_path")
    df = data.get("excel")
    root = data.get("xml")

    # 🔹 Preload all JSON data for duplication checks (independent of mode)
    missing_naming_variants = load_missing_naming_variants(paths["missing_naming_variants_json"])
    collocation_data = load_collocations_json(paths["collocations_json"])
    categorized_entries = load_json_annotations(paths["categorization_json"])

    # Defaults for optional tracking
    previous_namings = []
    previous_collocations = []
    previous_categorized_entries = []

    # Optional: lemma support for categorization
    lemma_normalization = None
    ignored_lemmas = None
    lemma_categories = None

    mode = config_data.get("modus", "collect")

    if mode == "analyze":
        run_analysis_menu(config_data, paths, data, book_name)
        return  # skip rest of collection logic

    # 🔹 6. User-controlled analysis paths
    check_namings = config_data["check_namings"]
    fill_collocations = config_data["fill_collocations"]
    do_categorization = config_data["do_categorization"]

    # 🔹 7. Load previous data depending on paths
    if check_namings:
        previous_verse = namings_last_verse
        active_last_verse = namings_last_verse
        previous_namings = missing_naming_variants.copy()

    elif fill_collocations:
        previous_verse = collocations_last_verse
        active_last_verse = collocations_last_verse
        previous_collocations = collocation_data.copy()

    elif do_categorization:
        previous_verse = categorization_last_verse
        active_last_verse = categorization_last_verse
        previous_categorized_entries = categorized_entries.copy()
        lemma_normalization = load_lemma_normalization(paths["lemma_normalization_json"])
        ignored_lemmas = load_ignored_lemmas(paths["ignored_lemmas_json"])
        lemma_categories = load_lemma_categories(paths["lemma_categories_json"])

    else:
        previous_verse = 0
        active_last_verse = 0

    # 🔹 8. Process TEI and run requested analysis steps
    missing_naming_variants, collocation_data, categorized_entries = run_data_collection(
        df=df,
        root=root,
        naming_dict=naming_dict,
        last_verse=active_last_verse,
        paths=paths,
        missing_naming_variants=missing_naming_variants,
        collocation_data=collocation_data,
        check_namings=check_namings,
        perform_collocations=fill_collocations,
        perform_categorization=do_categorization,
        lemma_normalization=lemma_normalization if do_categorization else None,
        ignored_lemmas=ignored_lemmas if do_categorization else None,
        lemma_categories=lemma_categories if do_categorization else None,
        categorized_entries=categorized_entries if do_categorization else None
    )

    # 🔹 9. Final save
    save_progress(
        missing_naming_variants=missing_naming_variants,
        last_processed_verse=active_last_verse,
        paths=paths,
        previous_verse=previous_verse,
        previous_namings=previous_namings,
        collocation_data=collocation_data,
        previous_collocations=previous_collocations,
        categorized_entries=categorized_entries,
        previous_categorized_entries=previous_categorized_entries,
        check_namings=check_namings,
        perform_collocations=fill_collocations,
        perform_categorization=do_categorization
    )

    # 🔹 10. Optional export
    export = ask_user_choice("Do you want to export all results? (y/n): ", ["y", "n"]) == "y"
    if export:
        paths["original_excel"] = data["excel_path"]
        options = {
            "benennungen": check_namings,
            "kollokationen": fill_collocations,
            "kategorisierung": do_categorization
        }
        export_all_data_to_new_excel(book_name, paths, options)

    # 🔹 11. Optional analysis after data collection
    analyze_after = ask_user_choice("Do you want to run an analysis now? (y/n): ", ["y", "n"])
    if analyze_after == "y":
        run_analysis_menu(config_data, paths, data, book_name)


if __name__ == "__main__":
    main()
