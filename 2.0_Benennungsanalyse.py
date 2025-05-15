import tkinter as tk
from tkinter import filedialog

import pandas as pd
import re
import json
import os
import shutil

import xml.etree.ElementTree as ET
from xml.etree.ElementTree import Element

import copy
from copy import copy

from typing import Union

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

DatenTyp = dict[str, Union[pd.DataFrame, Element, str, None]]
tei_ns = {'tei': 'http://www.tei-c.org/ns/1.0'}

def initialisiere_projekt():
    """
    Fragt den Benutzer nach dem Buchnamen, legt projektbezogene JSON-Pfade an
    und lädt ggf. vorhandene Fortschrittsdaten und fehlende Benennungen.
    Gibt ein Tupel zurück: (buchname, fehlende_benennungen, letzter_bearbeiteter_vers, pfade_dict)
    """

    buchname = input("Welches Buch bearbeiten wir heute? (z. B. Eneasroman): ").strip()

    # Verzeichnis anlegen
    os.makedirs("data", exist_ok=True)

    # Pfade vorbereiten (buchspezifisch)
    benennungen_json_path = os.path.join("data", f"fehlende_benennungen_{buchname}.json")
    progress_json_path = os.path.join("data", f"progress_{buchname}.json")
    kollokationen_json_path = os.path.join("data", f"kollokationen_{buchname}.json")
    kategorisierung_json_path = os.path.join("data", f"kategorisierung_{buchname}.json")
    lemma_normalisierung_json_path = os.path.join("data", f"lemma_normalisierung_{buchname}.json")
    ignorierte_lemmas_json_path = os.path.join("data", f"ignorierte_lemmas_{buchname}.json")
    lemma_kategorien_json_path = os.path.join("data", f"lemma_kategorien_{buchname}.json")

    paths = {
        "benennungen_json": benennungen_json_path,
        "progress_json": progress_json_path,
        "kollokationen_json": kollokationen_json_path,
        "kategorisierung_json": kategorisierung_json_path,
        "lemma_normalisierung_json": lemma_normalisierung_json_path,
        "ignorierte_lemmas_json": ignorierte_lemmas_json_path,
        "lemma_kategorien_json": lemma_kategorien_json_path
    }

    # Fortschritt laden oder auf 0 setzen
    letzter_bearbeiteter_vers = 0
    if os.path.exists(progress_json_path):
        with open(progress_json_path, "r", encoding="utf-8") as f:
            letzter_bearbeiteter_vers = json.load(f).get("letzter_vers", 0)

    # JSON-Dateien anlegen, falls sie fehlen
    initialisiere_dateien(paths)

    return buchname, letzter_bearbeiteter_vers, paths

def initialisiere_dateien(paths):
    """Legt die projektbezogenen JSON-Dateien an, falls sie noch nicht existieren."""

    def lege_an(pfad, inhalt):
        if not os.path.exists(pfad):
            with open(pfad, "w", encoding="utf-8") as f:
                json.dump(inhalt, f, indent=4, ensure_ascii=False)

    lege_an(paths["progress_json"], {"letzter_vers": 0})
    lege_an(paths["benennungen_json"], [])
    lege_an(paths["kollokationen_json"], [])
    lege_an(paths["kategorisierung_json"], [])

def lade_daten() -> DatenTyp:
    """Fragt interaktiv nach Excel- und TEI-Dateien, lädt sie bei Zustimmung und gibt sie gesammelt zurück."""
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    daten: DatenTyp = {"excel": None, "excel_pfad": None, "xml": None}

    # 1. Excel-Tabelle laden oder neu anlegen
    antwort_excel = input("Möchtest du eine Excel-Tabelle mit bereits erhobenen Benennungen laden? (j/n): ").strip().lower()
    if antwort_excel == "j":
        excel_pfad = filedialog.askopenfilename(
            title="Wähle die Excel-Datei mit den Benennungen",
            initialdir=os.getcwd(),
            filetypes=[("Excel-Dateien", "*.xlsx")]
        )
        if excel_pfad:
            while True:
                try:
                    daten["excel"] = pd.read_excel(excel_pfad)
                    daten["excel"] = pruefe_pflichtspalten(daten["excel"])
                    daten["excel_pfad"] = excel_pfad
                    print(f"✅ Excel-Datei geladen: {os.path.basename(excel_pfad)}")
                    break  # erfolgreich geladen, Schleife beenden
                except PermissionError:
                    print("❌ Die Excel-Datei ist aktuell geöffnet oder gesperrt.")
                    print("🔁 Bitte schließe die Datei und wähle sie anschließend erneut aus.")
                    excel_pfad = filedialog.askopenfilename(
                        title="Wähle die Excel-Datei mit den Benennungen erneut",
                        initialdir=os.getcwd(),
                        filetypes=[("Excel-Dateien", "*.xlsx")]
                    )
                    if not excel_pfad:
                        print("⚠️ Keine Datei ausgewählt – Abbruch.")
                        break
                except Exception as e:
                    print(f"❌ Fehler beim Laden der Excel-Datei: {e}")
                    break
        else:
            print("⚠️ Keine Excel-Datei ausgewählt.")

    elif antwort_excel == "n":
        neue_excel = input("Möchtest du stattdessen eine neue Excel-Datei anlegen? (j/n): ").strip().lower()
        if neue_excel == "j":
            speicherpfad = filedialog.asksaveasfilename(
                title="Speicherort für neue Excel-Datei wählen",
                defaultextension=".xlsx",
                initialdir=os.getcwd(),
                filetypes=[("Excel-Dateien", "*.xlsx")]
            )
            if speicherpfad:
                try:
                    vorlage_pfad = os.path.join(os.getcwd(), "vorlage_excel.xlsx")
                    wb = load_workbook(vorlage_pfad)
                    wb.save(speicherpfad)
                    daten["excel"] = pd.read_excel(speicherpfad)
                    daten["excel"] = pruefe_pflichtspalten(daten["excel"])
                    daten["excel_pfad"] = speicherpfad
                    print(f"✅ Neue Excel-Datei angelegt: {os.path.basename(speicherpfad)}")
                except Exception as e:
                    print(f"❌ Fehler beim Erstellen der neuen Datei: {e}")
            else:
                print("⚠️ Kein Speicherort ausgewählt.")

    # 2. TEI-XML-Datei laden?
    antwort_xml = input("Möchtest du die dazugehörige TEI-Datei laden? (j/n): ").strip().lower()
    if antwort_xml == "j":
        xml_pfad = filedialog.askopenfilename(
            title="Wähle die TEI-XML-Datei",
            initialdir=os.getcwd(),
            filetypes=[("XML-Dateien", "*.xml")]
        )
        if xml_pfad:
            try:
                tree = ET.parse(xml_pfad)
                root = tree.getroot()
                root = normalisiere_tei_text(root)
                daten["xml"] = root
                print(f"✅ XML-Datei geladen: {os.path.basename(xml_pfad)}")
            except Exception as e:
                print(f"❌ Fehler beim Laden der XML-Datei: {e}")
        else:
            print("⚠️ Keine XML-Datei ausgewählt.")

    return daten

def sortierte_eintraege(liste: list) -> list:
    """
    Gibt eine sortierte Kopie der Einträge zurück – nach Vers und Benennungswert.
    Damit können zwei Listen stabil miteinander verglichen werden.
    """
    return sorted(
        copy.deepcopy(liste),
        key=lambda x: (
            x.get("Vers", 0),
            x.get("Eigennennung") or x.get("Bezeichnung") or x.get("Erzähler") or ""
        )
    )

def pruefe_pflichtspalten(df: pd.DataFrame) -> pd.DataFrame:
    """
    Prüft, ob alle Pflichtspalten im DataFrame vorhanden sind.
    Bei fehlenden Spalten wird nachgefragt, ob sie automatisch ergänzt werden sollen.
    Gibt den ggf. erweiterten DataFrame zurück.
    """
    pflichtspalten = [
        "benannte figur",
        "vers",
        "eigennennung",
        "nennende figur",
        "bezeichnung",
        "erzähler",
        "kollokationen"
    ]

    aktuelle_spalten_lower = [sp.lower() for sp in df.columns]
    fehlende_spalten = [sp for sp in pflichtspalten if sp not in aktuelle_spalten_lower]

    if not fehlende_spalten:
        print("✅ Alle Pflichtspalten sind vorhanden.")
        return df

    print("⚠️ Folgende Pflichtspalten fehlen:")
    for spalte in fehlende_spalten:
        print(f"   – {spalte}")

    for spalte in fehlende_spalten:
        antwort = input(f"Möchtest du die Spalte „{spalte}“ automatisch ergänzen? (j/n): ").strip().lower()
        if antwort == "j":
            df[spalte] = ""
            print(f"➕ Spalte „{spalte}“ ergänzt (leer).")
        else:
            print(f"⚠️ Spalte „{spalte}“ bleibt fehlend.")

    return df

def normalisiere_text(text):
    """Normalisiert einen gegebenen Text nach festgelegten Regeln."""
    ersetzungen = {
        'æ': 'ae', 'œ': 'oe',
        'é': 'e', 'è': 'e', 'ë': 'e', 'á': 'a', 'à': 'a',
        'û': 'u', 'î': 'i', 'â': 'a', 'ô': 'o', 'ê': 'e',
        'ü': 'u', 'ö': 'o', 'ä': 'a',
        'ß': 'ss',
        'iu': 'ie', 'üe': 'ue'
    }

    if not text:
        return ""

    text = text.lower()
    for alt, neu in ersetzungen.items():
        text = text.replace(alt, neu)

    text = re.sub(r'\bv\b', 'f', text)  # Ersetze 'v' am Wortanfang durch 'f'
    text = re.sub(r'\s+', ' ', text)   # Mehrfache Leerzeichen zusammenfassen

    return text

def normalisiere_tei_text(root):
    """Normalisiert alle Texte innerhalb der TEI-Datei."""
    if root is None:
        return None

    normalisierte_verse = []
    for seg in root.findall('.//tei:seg', tei_ns):
        if seg.text:
            normalisierter_text = normalisiere_text(seg.text)
            seg.text = normalisierter_text
            normalisierte_verse.append(normalisierter_text)

    print("✅ TEI-Text wurde normalisiert.")

    return root

def speichere_fortschritt(
    fehlende_benennungen,
    letzter_bearbeiteter_vers,
    paths,
    vorheriger_vers=None,
    vorherige_benennungen=None,
    kollokationen_daten=None,
    vorherige_kollokationen=None,
    kategorisierte_eintraege = None,
    vorherige_kategorisierte_eintraege = None
):
    """
    Speichert Fortschritt, Benennungen und ggf. Kollokationen nur,
    wenn sich im Vergleich zum vorherigen Stand etwas geändert hat.
    """

    # 📌 Fortschritt speichern – nur wenn sich etwas geändert hat
    if vorheriger_vers is None or letzter_bearbeiteter_vers != vorheriger_vers:
        with open(paths["progress_json"], "w", encoding="utf-8") as f:
            json.dump({"letzter_vers": letzter_bearbeiteter_vers}, f, indent=4, ensure_ascii=False)

    # 📌 Benennungen speichern – nur wenn sich etwas geändert hat
    if vorherige_benennungen is None or sortierte_eintraege(fehlende_benennungen) != sortierte_eintraege(vorherige_benennungen):
        with open(paths["benennungen_json"], "w", encoding="utf-8") as f:
            json.dump(fehlende_benennungen, f, indent=4, ensure_ascii=False)

    # 📌 Kollokationen speichern – nur wenn übergeben und geändert
    if kollokationen_daten is not None:
        if vorherige_kollokationen is None or kollokationen_daten != vorherige_kollokationen:
            with open(paths["kollokationen_json"], "w", encoding="utf-8") as f:
                json.dump(kollokationen_daten, f, indent=4, ensure_ascii=False)

    # 📌 Kategorisierung speichern – nur wenn übergeben und geändert
    if kategorisierte_eintraege is not None:
        if vorherige_kategorisierte_eintraege is None or sortierte_eintraege(kategorisierte_eintraege) != sortierte_eintraege(vorherige_kategorisierte_eintraege):
            with open(paths["kategorisierung_json"], "w", encoding="utf-8") as f:
                json.dump(kategorisierte_eintraege, f, indent=4, ensure_ascii=False)


def lade_oder_erweitere_benennungen_dict():
    """
    Lädt oder erstellt ein zentrales Dictionary mit Figurenbenennungen aus Excel-Dateien.
    Rückgabe: dict mit Struktur {'Enthaltene Bücher': [...], 'Benennungen': {buch: [benennungen, ...]}}
    """

    os.makedirs("data", exist_ok=True)
    dict_path = os.path.join("data", "benennungen_dict.json")

    # Bestehendes Dict laden oder neues anlegen
    if os.path.exists(dict_path):
        with open(dict_path, "r", encoding="utf-8") as f:
            benennungen_dict = json.load(f)
        print(f"📚 Es wurde ein Dictionary gefunden.")
        buecher_liste = benennungen_dict.get("Enthaltene Bücher", [])
        if buecher_liste:
            print(f"👉 Enthaltene Bücher: {', '.join(buecher_liste)}")
        else:
            print("👉 Enthaltene Bücher: [leer]")
        erweitern = input("Möchtest du eine Datei ergänzen? (j/n): ").strip().lower()
    else:
        benennungen_dict = {"Enthaltene Bücher": [], "Benennungen": {}}
        print("❗ Es wurde noch kein Benennungs-Dictionary gefunden.")
        erweitern = "j"

    while erweitern == "j":
        print("📂 Bitte wähle eine Excel-Datei mit Benennungsdaten aus.")
        tk.Tk().withdraw()
        file_path = filedialog.askopenfilename(title="Excel-Datei auswählen", filetypes=[("Excel-Dateien", "*.xlsx")])
        if not file_path:
            print("⚠️ Keine Datei gewählt. Vorgang abgebrochen.")
            break

        buchname = input("Wie lautet der Name des Buchs (z. B. Eneasroman)? ").strip()

        try:
            df = pd.read_excel(file_path)
            relevante_spalten = ["Eigennennung", "Bezeichnung", "Erzähler"]
            benennungen = []

            for spalte in relevante_spalten:
                if spalte in df.columns:
                    benennungen.extend(df[spalte].dropna().tolist())

            # Doppelte entfernen und bereinigen
            benennungen = list(set(str(f).strip().lower() for f in benennungen if str(f).strip()))

        except Exception as e:
            print(f"❌ Fehler beim Einlesen der Datei: {e}")
            break

        benennungen_dict["Enthaltene Bücher"].append(buchname)
        benennungen_dict["Benennungen"][buchname] = benennungen
        print(f"✅ Buch '{buchname}' mit {len(benennungen)} Benennungen hinzugefügt.")

        erweitern = input("Möchtest du eine Datei ergänzen? (j/n): ").strip().lower()

    with open(dict_path, "w", encoding="utf-8") as f:
        json.dump(benennungen_dict, f, indent=4, ensure_ascii=False)
        print(f"💾 Aktuelles Dictionary unter: {dict_path}")

    return benennungen_dict

def durchsuche_tei_mit_dict(
    df,
    root,
    benennungen_dict,
    letzter_vers,
    paths,
    fehlende_benennungen,
    kollokationen_daten,
    pruefe_benennungen=True,
    fuehre_kollokationen_durch=False,
    fuehre_kategorisierung_durch=False,
    lemma_normalisierung=None,
    ignorierte_lemmas=None,
    lemma_kategorien=None,
    kategorisierte_eintraege=None):

    """
    Durchläuft den TEI-Text ab gespeichertem Vers und führt die gewählten Prüfungen aus.
    """

    if root is None or df is None or benennungen_dict is None:
        print("⚠️ Ungültige Eingaben – Abbruch.")
        return fehlende_benennungen

    verse = root.findall('.//tei:l', tei_ns)
    if not verse:
        print("⚠️ Keine Verse gefunden.")
        return fehlende_benennungen

    start_index = next((i for i, line in enumerate(verse) if int(line.get("n")) >= letzter_vers), 0)

    print(f"🔁 Starte Durchlauf ab Vers {int(verse[start_index].get('n'))} (Index {start_index})")

    for line in verse[start_index:]:
        vers_nr = int(line.get("n"))

        verse_text = ' '.join([seg.text for seg in line.findall(".//tei:seg", tei_ns) if seg.text])
        normalized_verse = normalisiere_text(verse_text)

        if pruefe_benennungen:
            fehlende_benennungen = pruefe_und_ergaenze_benennungen(
                vers_nr, verse_text, normalized_verse, df, benennungen_dict, fehlende_benennungen, root, paths
            )

        if fuehre_kollokationen_durch:
            pruefe_und_ergaenze_kollokationen(
                vers_nr, df, kollokationen_daten, root, paths
            )

        if fuehre_kategorisierung_durch:
            eintraege = df[df["Vers"] == vers_nr].to_dict(orient="records")
            for entry in eintraege:
                annotiert = lemmatisiere_und_kategorisiere_eintrag(
                    entry,
                    lemma_normalisierung,
                    ignorierte_lemmas,
                    lemma_kategorien
                )
                if annotiert:
                    kategorisierte_eintraege.append(annotiert)

        # Fortschritt speichern
        speichere_fortschritt(
            fehlende_benennungen=fehlende_benennungen,
            letzter_bearbeiteter_vers=vers_nr,
            paths=paths
        )

    return fehlende_benennungen


def pruefe_und_ergaenze_benennungen(
    vers_nr: int,
    verse_text: str,
    normalized_verse: str,
    df: pd.DataFrame,
    benennungen_dict: dict,
    fehlende_benennungen: list,
    root: Element,
    paths: dict
) -> list:
    """
    Prüft, ob eine Benennung aus dem globalen Dict im aktuellen Vers vorkommt,
    aber nicht in Excel oder in bereits bestätigten/abgelehnten Benennungen.
    Bei Treffer: Interaktive Ergänzung + Speicherung.
    """

    # 1. Benennungen aus Excel für diesen Vers extrahieren
    vorhandene_benennungen = set()
    if "Vers" in df.columns:
        df_vers = df[df["Vers"] == vers_nr]
        for spalte in ["Eigennennung", "Bezeichnung", "Erzähler"]:
            if spalte in df_vers.columns:
                werte = df_vers[spalte].dropna().tolist()
                vorhandene_benennungen.update(
                    normalisiere_text(str(wert).strip()) for wert in werte if str(wert).strip()
                )

    # 2. Benennungen aus Dict extrahieren & normalisieren
    dict_benennungen = set()
    for buchliste in benennungen_dict.get("Benennungen", {}).values():
        dict_benennungen.update(
            normalisiere_text(name.strip()) for name in buchliste if name.strip()
        )

    # 3. Fundprüfung & Benutzerinteraktion
    for benennung in dict_benennungen:
        if not benennung:
            continue

        # überspringen, wenn bereits in Excel oder JSON behandelt
        if any(benennung in eintrag for eintrag in vorhandene_benennungen) or any(
            vers_nr == eintrag.get("Vers") and
            normalisiere_text(benennung) == normalisiere_text(
                eintrag.get("Eigennennung") or eintrag.get("Bezeichnung") or eintrag.get("Erzähler") or ""
            )
            for eintrag in fehlende_benennungen
        ):
            continue

        if not re.search(rf'\b{re.escape(benennung)}\b', normalized_verse):
            continue

        print("\n" + "-" * 60)
        print(f"❗ Neue Benennung gefunden, die nicht in der Excel-Datei existiert!")
        print(f"🔍 Gefundene Benennung: \"{benennung}\"")

        # 📖 Kontext anzeigen
        prev_line = root.find(f'.//tei:l[@n="{vers_nr - 1}"]', tei_ns)
        if prev_line is not None:
            prev_text = ' '.join([seg.text for seg in prev_line.findall('.//tei:seg', tei_ns) if seg.text])
            print(f"📖 Vorheriger Vers ({vers_nr - 1}): {prev_text}")

        highlighted = verse_text.replace(benennung, f"\033[1m\033[93m{benennung}\033[0m")
        print(f"📖 Vers ({vers_nr}): {highlighted}")

        next_line = root.find(f'.//tei:l[@n="{vers_nr + 1}"]', tei_ns)
        if next_line is not None:
            next_text = ' '.join([seg.text for seg in next_line.findall('.//tei:seg', tei_ns) if seg.text])
            print(f"📖 Nächster Vers ({vers_nr + 1}): {next_text}")

        # 🧍 Benutzerabfrage
        confirm = input("Ist dies eine fehlende Benennung? (j/n): ").strip().lower()
        if confirm == "n":
            fehlende_benennungen.append({
                "Vers": vers_nr,
                "Eigennennung": benennung,
                "Nennende Figur": "",
                "Bezeichnung": "",
                "Erzähler": "",
                "Status": "abgelehnt"
            })
            speichere_fortschritt(fehlende_benennungen, vers_nr, paths)
            print("✅ Ablehnung gespeichert.")
            continue

        erweitern = input("💡 Möglicherweise ist dies eine mehrteilige Benennung. Erweitern? (j/n): ").strip().lower()
        if erweitern == "j":
            benennung = input("✍ Gib die vollständige Benennung ein: ").strip()

        print("Bitte wähle die richtige Kategorie:")
        print("[1] Eigennennung")
        print("[2] Bezeichnung")
        print("[3] Erzähler")
        print("[4] Überspringen")

        choice = input("👉 Deine Auswahl: ").strip()
        if choice == "4":
            continue

        benannte_figur = input("Gib die \"Benannte Figur\" ein: ").strip()
        nennende_figur = ""
        if choice == "2":
            nennende_figur = input("Gib die \"Nennende Figur\" ein: ").strip()

        eintrag = {
            "Benannte Figur": benannte_figur,
            "Vers": vers_nr,
            "Eigennennung": benennung if choice == "1" else "",
            "Nennende Figur": nennende_figur,
            "Bezeichnung": benennung if choice == "2" else "",
            "Erzähler": benennung if choice == "3" else "",
            "Status": "bestätigt"
        }

        # 📌 Optionale Kollokation
        will_kollokation = input("📌 Möchtest du eine Kollokation (Kontextstellen) hinzufügen? (j/n): ").strip().lower()
        if will_kollokation == "j":
            print("\n📖 Erweiterter Kontext (1–13):")
            vers_kontext = {}
            nummer = 1

            for i in range(6, 0, -1):
                zeile = root.find(f'.//tei:l[@n="{vers_nr - i}"]', tei_ns)
                if zeile is not None:
                    text = ' '.join([seg.text for seg in zeile.findall('.//tei:seg', tei_ns) if seg.text])
                    vers_kontext[nummer] = text
                    print(f"[{nummer}] {text}")
                    nummer += 1

            vers_kontext[nummer] = verse_text
            print(f"[{nummer}] {verse_text}")
            nummer += 1

            for i in range(1, 7):
                zeile = root.find(f'.//tei:l[@n="{vers_nr + i}"]', tei_ns)
                if zeile is not None:
                    text = ' '.join([seg.text for seg in zeile.findall('.//tei:seg', tei_ns) if seg.text])
                    vers_kontext[nummer] = text
                    print(f"[{nummer}] {text}")
                    nummer += 1

            auswahl = input("\n👉 Bitte gib die Versnummer(n) ein (z.B. '5-7' oder '6'): ").strip()
            ausgewaehlt = []

            try:
                if "-" in auswahl:
                    start, ende = map(int, auswahl.split("-"))
                    ausgewaehlt = [vers_kontext[i] for i in range(start, ende + 1) if i in vers_kontext]
                else:
                    idx = int(auswahl)
                    ausgewaehlt = [vers_kontext[idx]]
            except (ValueError, KeyError):
                print("⚠️ Ungültige Eingabe – keine Kollokation gespeichert.")

            if ausgewaehlt:
                eintrag["Kollokation"] = ' / '.join(ausgewaehlt)

        fehlende_benennungen.append(eintrag)
        speichere_fortschritt(fehlende_benennungen, vers_nr, paths)
        print("✅ Eintrag gespeichert.")

    return fehlende_benennungen

def lade_fehlende_benennungen(pfad: str) -> list:
    """
    Lädt fehlende oder bestätigte Benennungen aus einer JSON-Datei.
    Gibt eine leere Liste zurück, wenn die Datei nicht existiert oder fehlerhaft ist.
    """
    if os.path.exists(pfad):
        try:
            with open(pfad, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, FileNotFoundError):
            print("⚠️ Fehler beim Laden der JSON-Datei – leere Liste wird verwendet.")
            return []
    else:
        return []

def hole_verskontext(vers_nr, root_tei):
    """Holt die umgebenden 6 Verse aus der TEI-Datei, nummeriert sie von 1–13."""
    kontext = []
    vers_liste = []

    for i in range(-6, 7):
        vers_id = str(vers_nr + i)  # explizit String!
        line = root_tei.find(f'.//tei:l[@n="{vers_id}"]', tei_ns)

        if line is not None:
            text = normalisiere_text(' '.join([
                seg.text for seg in line.findall('.//tei:seg', tei_ns) if seg.text
            ]))
            vers_liste.append(text)

    for i, vers in enumerate(vers_liste, start=1):
        kontext.append((i, vers))

    return kontext

def bereinige_zellenwert(value):
    if pd.isna(value) or value is None:
        return ""
    return normalisiere_text(str(value).strip())

def pruefe_und_ergaenze_kollokationen(vers_nr, df, kollokationen_daten, root, paths):

    """Prüft, ob eine Kollokation ergänzt werden soll – falls ja, ruft UI auf."""

    zeilen = df[df["Vers"] == vers_nr]
    if zeilen.empty:
        return None

    zeile = zeilen.iloc[0]

    if pd.notna(zeile.get("Kollokationen")) and str(zeile["Kollokationen"]).strip() != "":
        return None

    if any(eintrag["Vers"] == vers_nr for eintrag in kollokationen_daten):
        return None

    benennung = bereinige_zellenwert(zeile.get("Eigennennung")) \
                or bereinige_zellenwert(zeile.get("Bezeichnung")) \
                or bereinige_zellenwert(zeile.get("Erzähler"))

    benannte_figur = bereinige_zellenwert(zeile.get("Benannte Figur"))

    kontext = hole_verskontext(vers_nr, root)

    kollokationen = frage_nach_kollokationen(vers_nr, benannte_figur, benennung, kontext)

    kollokationen_daten.append({
        "Vers": vers_nr,
        "Kollokationen": kollokationen
    })

    # 📝 Sofortige Zwischenspeicherung nach erfolgreicher Auswahl
    with open(paths["kollokationen_json"], "w", encoding="utf-8") as f:
        json.dump(kollokationen_daten, f, indent=4, ensure_ascii=False)

    return True

def frage_nach_kollokationen(vers_nr, benannte_figur, benennung, kontext):
    """Zeigt den Kontext eines Verses und fragt interaktiv nach relevanten Kollokationen."""

    print(f"\n🟡 Leere Kollokationen in Vers {vers_nr} erkannt!")
    if benannte_figur or benennung:
        print(f"👤 {benannte_figur}: {benennung}\n")

    for nummer, text in kontext:
        if benennung:
            # Benennung hervorheben
            hervorgehoben = text.replace(str(benennung), f"\033[1;33m{benennung}\033[0m")
        else:
            hervorgehoben = text
        print(f"{nummer}. {hervorgehoben}")

    # Nutzereingabe
    eingabe = input("\n👉 Bitte gib die Nummer(n) der relevanten Verse ein (z. B. '5' oder '5-7'): ")

    # wird unten im return verwendet
    ausgewaehlt = []

    try:
        if "-" in eingabe:
            start, ende = map(int, eingabe.split("-"))
            ausgewaehlt = [text for nummer, text in kontext if start <= nummer <= ende]
        else:
            nummer = int(eingabe)
            ausgewaehlt = [text for num, text in kontext if num == nummer]
    except (ValueError, StopIteration):
        print("⚠️ Ungültige Eingabe. Bitte eine Zahl oder einen Bereich eingeben.")
        return ""

    return " / ".join(ausgewaehlt)

def lade_kollokationen_json(pfad_zur_datei):
    """Lädt vorhandene Kollokationen aus einer JSON-Datei – oder gibt leere Liste zurück."""
    if os.path.exists(pfad_zur_datei):
        with open(pfad_zur_datei, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []

def lade_ignorierte_lemmas(path="ignorierte_lemmas.json"):
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
            return set(data) if isinstance(data, list) else set(data.keys())
    return set()

def lade_lemma_kategorien(path="lemma_kategorien.json"):
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def lade_json_annotationen(path):
    if os.path.exists(path):
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []

def speichere_json_annotationen(path, annotations):
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(annotations, f, ensure_ascii=False, indent=2)

def lemmatisiere_und_kategorisiere_eintrag(entry, lemma_normalisierung, ignorierte_lemmas=None, lemma_kategorien=None):

    if lemma_normalisierung is None:
        lemma_normalisierung = {}

    if ignorierte_lemmas is None:
        ignorierte_lemmas = lade_ignorierte_lemmas()

    if lemma_kategorien is None:
        lemma_kategorien = lade_lemma_kategorien()

    text = entry.get("Erzähler") or entry.get("Bezeichnung") or entry.get("Eigennennung")
    if not text:
        print("⚠ Kein Text zum Annotieren vorhanden – Eintrag wird übersprungen.\n")
        return None

    print("\n" + "=" * 60)
    print(f"▶ Vers: {entry.get('Vers')}")
    print(f"▶ Benannte Figur: {entry.get('Benannte Figur')}")
    typ = "Erzähler" if entry.get("Erzähler") else ("Bezeichnung" if entry.get("Bezeichnung") else "Eigennennung")
    print(f"▶ Typ: {typ}")
    print(f"\n▶ Originaltext: {text}")

    if ignorierte_lemmas is None:
        ignorierte_lemmas = lade_ignorierte_lemmas()
    if lemma_kategorien is None:
        lemma_kategorien = lade_lemma_kategorien()

    tokens = zerlege_in_tokens(text.lower())
    fehlende = [t for t in tokens if t not in lemma_normalisierung]

    if fehlende:
        print("\n▶ Lemmata bitte ergänzen (getrennt durch Komma):")
        user_input = input("> ").strip()
        neue_lemmata = [l.strip() for l in user_input.split(",") if l.strip()]
        if len(neue_lemmata) != len(fehlende):
            print("⚠ Anzahl der eingegebenen Lemmata stimmt nicht mit der Anzahl der unbekannten Tokens überein. Vorgang abgebrochen.\n")
            return None
        for token, lemma in zip(fehlende, neue_lemmata):
            lemma_normalisierung[token] = lemma
        speichere_lemma_normalisierung(lemma_normalisierung)

    lemmata = [lemma_normalisierung.get(t, t) for t in tokens]

    print(f"\n▶ Lemma: {', '.join(lemmata)}\n")

    bezeichnungen = []
    epitheta = []
    history = []

    i = 0
    while i < len(lemmata):
        lemma = lemmata[i]
        if lemma in ignorierte_lemmas:
            i += 1
            continue

        vorgabe = f"[{lemma_kategorien.get(lemma, '')}]" if lemma in lemma_kategorien else ""
        print(f"{lemma:<12} → {vorgabe} ", end="")
        user_input = input().strip()

        if user_input == "<":
            if i == 0 or not history:
                print("↩️  Bereits am Anfang – Rücksprung nicht möglich.")
                continue

            i -= 1
            last_action = history.pop()

            if last_action["type"] == "a":
                bezeichnungen.pop()
            elif last_action["type"] == "e":
                epitheta.pop()
            elif last_action["type"] == "ignore":
                ignorierte_lemmas.discard(last_action["lemma"])
                speichere_ignorierte_lemmas(ignorierte_lemmas)
            elif last_action["type"] == "override":
                del lemma_kategorien[last_action["lemma"]]
                speichere_lemma_kategorien(lemma_kategorien)
            continue

        if user_input == "" and vorgabe:
            if vorgabe == "[a]":
                bezeichnungen.append(lemma)
                history.append({"type": "a", "lemma": lemma})
            elif vorgabe == "[e]":
                epitheta.append(lemma)
                history.append({"type": "e", "lemma": lemma})
            i += 1
            continue

        if user_input == "":
            ignorierte_lemmas.add(lemma)
            speichere_ignorierte_lemmas(ignorierte_lemmas)
            print(f"ℹ️ Lemma „{lemma}“ zur Ignorierliste hinzugefügt.")
            history.append({"type": "ignore", "lemma": lemma})
            i += 1
            continue

        if user_input in ("a", "e"):
            if user_input == "a":
                bezeichnungen.append(lemma)
            else:
                epitheta.append(lemma)
            lemma_kategorien[lemma] = user_input
            speichere_lemma_kategorien(lemma_kategorien)
            history.append({"type": user_input, "lemma": lemma})
            i += 1
            continue

        korrektur = user_input
        kat = ""
        while kat not in ("a", "e"):
            kat = input(f'Definiere die Kategorie für „{korrektur}“ [a/e]: ').strip().lower()

        if kat == "a":
            bezeichnungen.append(korrektur)
        else:
            epitheta.append(korrektur)

        lemma_kategorien[korrektur] = kat
        speichere_lemma_kategorien(lemma_kategorien)

        history.append({
            "type": "override",
            "lemma": korrektur
        })
        i += 1

    if not bezeichnungen and not epitheta:
        print("⚠ Kein Eintrag – bitte prüfe und bestätige erneut.")
        confirm = input("Eintrag wirklich überspringen? [j = ja / n = nein]: ").strip().lower()
        if confirm == "j":
            print("⏭ Eintrag wurde übersprungen.\n")
            return None
        else:
            return lemmatisiere_und_kategorisiere_eintrag(entry, lemma_normalisierung, ignorierte_lemmas, lemma_kategorien)

    print("✅ Eintrag automatisch gespeichert.\n")
    return {
        **entry,
        "Bezeichnung 1": bezeichnungen[0] if len(bezeichnungen) > 0 else "",
        "Bezeichnung 2": bezeichnungen[1] if len(bezeichnungen) > 1 else "",
        "Bezeichnung 3": bezeichnungen[2] if len(bezeichnungen) > 2 else "",
        "Bezeichnung 4": bezeichnungen[3] if len(bezeichnungen) > 3 else "",
        "Epitheta 1": epitheta[0] if len(epitheta) > 0 else "",
        "Epitheta 2": epitheta[1] if len(epitheta) > 1 else "",
        "Epitheta 3": epitheta[2] if len(epitheta) > 2 else "",
        "Epitheta 4": epitheta[3] if len(epitheta) > 3 else "",
        "Epitheta 5": epitheta[4] if len(epitheta) > 4 else ""
    }

def zerlege_in_tokens(text):
    return re.findall(r'\w+|[^\w\s]', text, re.UNICODE)

def lade_lemma_normalisierung(path="lemma_normalisierung.json"):
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def speichere_lemma_normalisierung(data, path="lemma_normalisierung.json"):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def speichere_ignorierte_lemmas(data, path="ignorierte_lemmas.json"):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(sorted(data), f, ensure_ascii=False, indent=2)

def speichere_lemma_kategorien(data, path="lemma_kategorien.json"):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def exportiere_alle_daten_in_neue_excel(paths, options):
    """
    Integriert bestätigte Benennungen, ergänzt Kollokationen und erstellt ein lemmatisiertes Arbeitsblatt.

    :param paths: Dictionary mit Dateipfaden (original_excel, json_benennungen, json_kollokationen, json_kategorisierung)
    :param options: Steuerung der Verarbeitungsschritte über Boolean-Werte: benennungen, kollokationen, kategorisierung
    """
    
    # Zielpfad erzeugen mit "_final" im Dateinamen
    original_name = os.path.basename(paths["original_excel"])
    zielname = original_name.replace(".xlsx", "_final.xlsx")

    # 🔧 Verzeichnis sicherstellen
    zielverzeichnis = "/mnt/endproduct"
    os.makedirs(zielverzeichnis, exist_ok=True)

    zielpfad = os.path.join(zielverzeichnis, str(zielname))

    # Excel-Datei kopieren
    shutil.copy(paths["original_excel"], zielpfad)

    wb = openpyxl.load_workbook(zielpfad)
    sheet = wb["Gesamt"]

    if options.get("benennungen", False):
        fuege_benennungen_ein(sheet, paths["json_benennungen"])

    if options.get("kollokationen", False):
        aktualisiere_kollokationen(sheet, paths["json_kollokationen"])

    if options.get("kategorisierung", False):
        erzeuge_blatt_mit_kategorisierten_lemmata(wb, sheet, paths["json_kategorisierung"])

    wb.save(zielpfad)
    print(f"✅ Export abgeschlossen: {zielpfad}")

def hole_formatvorlage(sheet, spalten_index):
    """
    Gibt Formatvorlage (Font, Alignment, Border, Number Format) der ersten befüllten Zelle in der Spalte zurück.
    """
    for zeile in range(2, sheet.max_row + 1):
        zelle = sheet.cell(row=zeile, column=spalten_index)
        if zelle.value:
            if zelle.has_style:
                return copy(zelle.font), copy(zelle.alignment), copy(zelle.border), zelle.number_format
    return None, None, None, None

def fuege_benennungen_ein(sheet, json_path):
    """
    Fügt bestätigte Benennungen in das Blatt 'Gesamt' ein.
    Formatierung pro Spalte wird über 'hole_formatvorlage()' übernommen.
    Neue Zeilen werden farblich hervorgehoben.
    """
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    confirmed_entries = [entry for entry in data if entry.get("Status") == "bestätigt"]
    if not confirmed_entries:
        print("ℹ️ Keine bestätigten Benennungen zum Einfügen.")
        return

    last_line = sheet.max_row + 1
    fill_color = PatternFill(start_color="4BACC6", end_color="4BACC6", fill_type="solid")

    for entry in confirmed_entries:
        new_line = [
            entry.get("Benannte Figur", ""),
            entry.get("Vers", ""),
            entry.get("Eigennennung", ""),
            entry.get("Nennende Figur", ""),
            entry.get("Bezeichnung", ""),
            entry.get("Erzähler", ""),
            entry.get("Kollokation", "")
        ]

        for col_num, value in enumerate(new_line, start=1):
            cell = sheet.cell(row=last_line, column=col_num, value=value)

            font_tpl, alignment_tpl, border_tpl, number_format_tpl = hole_formatvorlage(sheet, col_num)
            if font_tpl:
                cell.font = font_tpl
                cell.alignment = alignment_tpl
                cell.border = border_tpl
                cell.number_format = number_format_tpl

            cell.fill = fill_color

        last_line += 1

    print("✅ Benennungen erfolgreich ergänzt.")

def aktualisiere_kollokationen(sheet, json_path):
    """
    Aktualisiert die Spalte 'Kollokationen' anhand der JSON-Daten.
    Formatierung wird aus der ersten befüllten Zelle übernommen (über hole_formatvorlage()).
    """
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    header = [cell.value for cell in sheet[1]]
    try:
        vers_col = header.index("Vers") + 1
        kollokationen_col = header.index("Kollokationen") + 1
    except ValueError:
        print("❌ Spalten 'Vers' oder 'Kollokationen' nicht gefunden!")
        return

    vers_to_rows = {}
    for row in range(2, sheet.max_row + 1):
        vers_value = sheet.cell(row=row, column=vers_col).value
        if vers_value is not None:
            vers_to_rows.setdefault(int(vers_value), []).append(row)

    font_tpl, alignment_tpl, border_tpl, number_format_tpl = hole_formatvorlage(sheet, kollokationen_col)

    updated_count = 0
    for entry in data:
        vers = entry["Vers"]
        new_value = entry["Kollokationen"]
        matching_rows = vers_to_rows.get(vers, [])
        for row in matching_rows:
            cell = sheet.cell(row=row, column=kollokationen_col, value=new_value)
            if font_tpl:
                cell.font = font_tpl
                cell.alignment = alignment_tpl
                cell.border = border_tpl
                cell.number_format = number_format_tpl
            updated_count += 1

    print(f"✅ {updated_count} Kollokationen erfolgreich aktualisiert.")

def erzeuge_blatt_mit_kategorisierten_lemmata(wb, original_sheet, json_path):
    """
    Erstellt ein neues Arbeitsblatt 'lemmatisiert' mit strukturierten Bezeichnungen und Epitheta
    basierend auf den Daten aus json_path. Formatierung erfolgt spaltenweise über hole_formatvorlage().
    """
    with open(json_path, "r", encoding="utf-8") as f:
        annotations = json.load(f)

    # Ursprungsblatt referenzieren
    ws_original = original_sheet

    # Vorhandenes Blatt 'lemmatisiert' löschen
    if 'lemmatisiert' in wb.sheetnames:
        del wb['lemmatisiert']

    # Kopie des Originals anlegen
    ws_new = wb.copy_worksheet(ws_original)
    ws_new.title = 'lemmatisiert'

    # Blattposition direkt hinter 'Gesamt'
    if 'Gesamt' in wb.sheetnames:
        idx = wb.sheetnames.index('Gesamt') + 1
        # 📌 Achtung: Zugriff auf protected member, da keine offizielle Alternative
        wb._sheets.insert(idx, wb._sheets.pop(wb.sheetnames.index('lemmatisiert')))

    headers = [
        "Benannte Figur", "Vers", "Eigennennung", "Nennende Figur", "Bezeichnung", "Erzähler",
        "Bezeichnung 1", "Bezeichnung 2", "Bezeichnung 3", "Bezeichnung 4",
        "Epitheta 1", "Epitheta 2", "Epitheta 3", "Epitheta 4", "Epitheta 5"
    ]

    df_new = pd.DataFrame(annotations)
    for col in headers:
        if col not in df_new.columns:
            df_new[col] = ""

            # Falls "Nennende Figur" fehlt → über "Vers" aus "Gesamt" zuordnen
            if col == "Nennende Figur" and 'Gesamt' in wb.sheetnames:
                ws_gesamt = wb['Gesamt']
                vers_to_nennende = {}
                header_row = [cell.value for cell in ws_gesamt[1]]
                vers_idx = header_row.index("Vers") + 1
                nennende_idx = header_row.index("Nennende Figur") + 1

                for row in ws_gesamt.iter_rows(min_row=2):
                    vers_val = row[vers_idx - 1].value
                    nennende_val = row[nennende_idx - 1].value
                    if vers_val is not None:
                        vers_to_nennende[vers_val] = nennende_val

                df_new["Nennende Figur"] = df_new["Vers"].map(vers_to_nennende).fillna("")

    df_new = df_new[headers]

    # Alte "Kollokationen"-Spalte entfernen
    for col in ws_new.iter_cols(min_row=1, max_row=1):
        if col[0].value == "Kollokationen":
            ws_new.delete_cols(col[0].column, 1)
            break

    # Redundante Spalten entfernen
    existing_headers = [cell.value for cell in ws_new[1]]
    redundant_cols = [i + 1 for i, h in enumerate(existing_headers) if h in headers]
    for idx in sorted(redundant_cols, reverse=True):
        ws_new.delete_cols(idx)

    # Einfügen neuer Spalten mit Formatierung
    for col_offset, header in enumerate(headers):
        col_index = col_offset + 1
        ws_new.cell(row=1, column=col_index, value=header)
        font_tpl, alignment_tpl, border_tpl, number_format_tpl = hole_formatvorlage(ws_original, 1)

        for row_idx, value in enumerate(df_new[header], start=2):
            cell = ws_new.cell(row=row_idx, column=col_index, value=value)
            if font_tpl:
                cell.font = font_tpl
                cell.alignment = alignment_tpl
                cell.border = border_tpl
                cell.number_format = number_format_tpl

    ws_new.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    print("✅ Arbeitsblatt 'lemmatisiert' erfolgreich erzeugt.")

def main():
    # 🔹 1. Initialisierung: Buchwahl, Pfade, letzter Vers
    buchname, letzter_bearbeiteter_vers, paths = initialisiere_projekt()

    # 🔹 3. Globales Benennungs-Dict laden (aus allen Büchern)
    benennungen_dict = lade_oder_erweitere_benennungen_dict()

    # 🔹 2. Daten laden: Excel & TEI-XML
    daten = lade_daten()
    paths["original_excel"] = daten["excel_pfad"]
    df = daten["excel"]
    root = daten["xml"]

    # 🔹 4. Vorherigen Vers merken
    vorheriger_vers = letzter_bearbeiteter_vers

    # 🔹 5. Initialisierung der Zwischenspeicher
    fehlende_benennungen = []
    kollokationen_daten = []
    vorherige_benennungen = []
    vorherige_kollokationen = []
    kategorisierte_eintraege = []
    vorherige_kategorisierte_eintraege = []

    # 🔹 6. Globale Steuerung der Analysepfade (Benennung, Kollokation, Kategorisierung)
    antwort_benennungen = input("Sollen Benennungen geprüft und ergänzt werden? (j/n): ").strip().lower() == "j"
    antwort_kollokationen = input("Sollen leere Kollokationen befüllt werden? (j/n): ").strip().lower() == "j"
    antwort_kategorisierung = input("Sollen die Benennungen lemmatisiert und kategorisiert werden? (j/n): ").strip().lower() == "j"

    # 🔹 7. Je nach Analysepfad: Daten gezielt laden
    if antwort_benennungen:
        fehlende_benennungen = lade_fehlende_benennungen(paths["benennungen_json"])
        vorherige_benennungen = fehlende_benennungen.copy()

    if antwort_kollokationen:
        kollokationen_daten = lade_kollokationen_json(paths["kollokationen_json"])
        vorherige_kollokationen = kollokationen_daten.copy()


    if antwort_kategorisierung:
        lemma_normalisierung = lade_lemma_normalisierung(paths["lemma_normalisierung_json"])
        ignorierte_lemmas = lade_ignorierte_lemmas(paths["ignorierte_lemmas_json"])
        lemma_kategorien = lade_lemma_kategorien(paths["lemma_kategorien_json"])


    # 🔹 8. TEI durchlaufen & gewählte Prüfungen durchführen
    fehlende_benennungen = durchsuche_tei_mit_dict(
        df=df,
        root=root,
        benennungen_dict=benennungen_dict,
        letzter_vers=letzter_bearbeiteter_vers,
        paths=paths,
        fehlende_benennungen=fehlende_benennungen,
        kollokationen_daten=kollokationen_daten,
        pruefe_benennungen=antwort_benennungen,
        fuehre_kollokationen_durch=antwort_kollokationen,
        fuehre_kategorisierung_durch=antwort_kategorisierung
    )

    # 🔹 9. Abschließende Sicherung
    speichere_fortschritt(
        fehlende_benennungen=fehlende_benennungen,
        letzter_bearbeiteter_vers=letzter_bearbeiteter_vers,
        paths=paths,
        vorheriger_vers=vorheriger_vers,
        vorherige_benennungen=vorherige_benennungen,
        kollokationen_daten=kollokationen_daten,
        vorherige_kollokationen=vorherige_kollokationen,
        kategorisierte_eintraege=kategorisierte_eintraege,
        vorherige_kategorisierte_eintraege=vorherige_kategorisierte_eintraege
    )

    # 🔹 10. Export?
    antwort_export = input("Möchtest du alle Ergebnisse exportieren? (j/n): ").strip().lower() == "j"
    if antwort_export:
        paths["original_excel"] = daten["excel_pfad"]
        options = {
            "benennungen": antwort_benennungen,
            "kollokationen": antwort_kollokationen,
            "kategorisierung": antwort_kategorisierung
        }
        exportiere_alle_daten_in_neue_excel(paths, options)


if __name__ == "__main__":
    main()
